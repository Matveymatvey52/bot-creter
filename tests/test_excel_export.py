"""features/excel_export.py tests — mirrors tests/test_sales_analytics.py /
tests/test_word_export_module.py's style: direct unit tests over the module's
pure functions plus a DB-isolation smoke test for fetch_export_rows().

Run with: pytest tests/test_excel_export.py -v
"""
from __future__ import annotations

import io
import unittest
import tempfile
from pathlib import Path

import aiosqlite
from openpyxl import load_workbook

import features.excel_export as excel_export


class BuildWorkbookTests(unittest.TestCase):
    def test_returns_nonempty_xlsx_bytes(self):
        data = excel_export.build_workbook(["Имя", "Сумма"], [["Иван", 1000]])
        self.assertIsInstance(data, bytes)
        self.assertGreater(len(data), 0)

    def test_workbook_is_valid_and_openable(self):
        data = excel_export.build_workbook(["A", "B"], [[1, 2], [3, 4]])
        wb = load_workbook(io.BytesIO(data))
        sheet = wb.active
        self.assertEqual(sheet["A1"].value, "A")
        self.assertEqual(sheet["B1"].value, "B")
        self.assertEqual(sheet["A2"].value, 1)
        self.assertEqual(sheet["B2"].value, 2)

    def test_header_styling_applied(self):
        data = excel_export.build_workbook(["Имя", "Сумма"], [["Иван", 1000]])
        wb = load_workbook(io.BytesIO(data))
        sheet = wb.active
        header_cell = sheet["A1"]
        self.assertTrue(header_cell.font.bold)
        self.assertEqual(header_cell.font.color.rgb[-6:], excel_export.HEADER_FONT_COLOR)
        self.assertEqual(header_cell.fill.start_color.rgb[-6:], excel_export.HEADER_COLOR)

    def test_data_rows_not_bold(self):
        data = excel_export.build_workbook(["A"], [[1]])
        wb = load_workbook(io.BytesIO(data))
        sheet = wb.active
        self.assertFalse(sheet["A2"].font.bold)

    def test_column_auto_width_computed(self):
        data = excel_export.build_workbook(
            ["Короткий", "Длинное_название_столбца_очень_длинное"],
            [["x", "a much much longer cell value than the header itself, really quite long"]],
        )
        wb = load_workbook(io.BytesIO(data))
        sheet = wb.active
        width_a = sheet.column_dimensions["A"].width
        width_b = sheet.column_dimensions["B"].width
        self.assertGreaterEqual(width_a, excel_export._MIN_COLUMN_WIDTH)
        self.assertGreater(width_b, width_a)
        self.assertLessEqual(width_b, excel_export._MAX_COLUMN_WIDTH)

    def test_freeze_panes_set(self):
        data = excel_export.build_workbook(["A"], [[1], [2]])
        wb = load_workbook(io.BytesIO(data))
        sheet = wb.active
        self.assertEqual(sheet.freeze_panes, "A2")

    def test_autofilter_enabled(self):
        data = excel_export.build_workbook(["A", "B"], [[1, 2], [3, 4]])
        wb = load_workbook(io.BytesIO(data))
        sheet = wb.active
        self.assertIsNotNone(sheet.auto_filter.ref)
        self.assertEqual(sheet.auto_filter.ref, "A1:B3")

    def test_max_rows_truncation(self):
        rows = [[i] for i in range(50)]
        data = excel_export.build_workbook(["A"], rows, max_rows=10)
        wb = load_workbook(io.BytesIO(data))
        sheet = wb.active
        # header + 10 data rows = 11
        self.assertEqual(sheet.max_row, 11)

    def test_empty_data_does_not_raise(self):
        data = excel_export.build_workbook(["A", "B"], [])
        wb = load_workbook(io.BytesIO(data))
        sheet = wb.active
        self.assertEqual(sheet["A1"].value, "A")
        self.assertEqual(sheet.max_row, 1)

    def test_empty_columns_and_rows_does_not_raise(self):
        data = excel_export.build_workbook([], [])
        self.assertIsInstance(data, bytes)
        self.assertGreater(len(data), 0)

    def test_status_column_conditional_formatting(self):
        data = excel_export.build_workbook(
            ["Название", "Статус"],
            [["Заказ 1", "done"], ["Заказ 2", "pending"], ["Заказ 3", "cancelled"]],
        )
        wb = load_workbook(io.BytesIO(data))
        sheet = wb.active
        # done -> green-ish, pending -> yellow-ish, cancelled -> red-ish
        self.assertEqual(sheet["B2"].fill.start_color.rgb[-6:], "C6EFCE")
        self.assertEqual(sheet["B3"].fill.start_color.rgb[-6:], "FFEB9C")
        self.assertEqual(sheet["B4"].fill.start_color.rgb[-6:], "FFC7CE")

    def test_non_status_column_no_conditional_fill(self):
        data = excel_export.build_workbook(["Имя"], [["done"]])
        wb = load_workbook(io.BytesIO(data))
        sheet = wb.active
        # "Имя" is not a status-like column name — value "done" must not get colored
        fill = sheet["A2"].fill
        self.assertIn(fill.fill_type, (None, "none"))

    def test_bytes_and_exotic_values_stringified(self):
        data = excel_export.build_workbook(["A"], [[b"\x00\x01"]])
        wb = load_workbook(io.BytesIO(data))
        sheet = wb.active
        self.assertIn("binary", str(sheet["A2"].value))

    def test_sheet_title_truncated_to_excel_limit(self):
        long_title = "x" * 50
        data = excel_export.build_workbook(["A"], [[1]], sheet_title=long_title)
        wb = load_workbook(io.BytesIO(data))
        self.assertLessEqual(len(wb.active.title), 31)


class FetchExportRowsTests(unittest.IsolatedAsyncioTestCase):
    async def asyncSetUp(self):
        self._tmp = tempfile.TemporaryDirectory()
        self.db_path = str(Path(self._tmp.name) / "bot_data.db")
        async with aiosqlite.connect(self.db_path) as db:
            await db.execute(
                "CREATE TABLE orders (id INTEGER PRIMARY KEY AUTOINCREMENT, customer TEXT, amount INTEGER, status TEXT)"
            )
            await db.execute("INSERT INTO orders (customer, amount, status) VALUES (?,?,?)", ("Иван", 1000, "done"))
            await db.execute("INSERT INTO orders (customer, amount, status) VALUES (?,?,?)", ("Пётр", 2000, "pending"))
            await db.commit()

    async def asyncTearDown(self):
        self._tmp.cleanup()

    async def test_no_hook_config_returns_none(self):
        result = await excel_export.fetch_export_rows(self.db_path, None)
        self.assertIsNone(result)

    async def test_invalid_table_identifier_returns_none(self):
        result = await excel_export.fetch_export_rows(self.db_path, {"table": "orders; DROP TABLE orders"})
        self.assertIsNone(result)

    async def test_stale_table_returns_none(self):
        result = await excel_export.fetch_export_rows(self.db_path, {"table": "nonexistent_table"})
        self.assertIsNone(result)

    async def test_fetches_columns_and_rows(self):
        result = await excel_export.fetch_export_rows(self.db_path, {"table": "orders"})
        self.assertIsNotNone(result)
        columns, rows = result
        self.assertEqual(columns, ["id", "customer", "amount", "status"])
        self.assertEqual(len(rows), 2)

    async def test_max_rows_cap_applied_at_fetch(self):
        async with aiosqlite.connect(self.db_path) as db:
            for i in range(20):
                await db.execute("INSERT INTO orders (customer, amount, status) VALUES (?,?,?)", (f"c{i}", i, "done"))
            await db.commit()
        result = await excel_export.fetch_export_rows(self.db_path, {"table": "orders"}, max_rows=5)
        self.assertIsNotNone(result)
        _columns, rows = result
        self.assertEqual(len(rows), 5)

    async def test_roundtrip_into_build_workbook(self):
        result = await excel_export.fetch_export_rows(self.db_path, {"table": "orders"})
        self.assertIsNotNone(result)
        columns, rows = result
        data = excel_export.build_workbook(columns, rows, sheet_title="orders")
        wb = load_workbook(io.BytesIO(data))
        sheet = wb.active
        self.assertEqual(sheet.max_row, 3)  # header + 2 rows


if __name__ == "__main__":
    unittest.main()
