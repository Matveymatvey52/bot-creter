# TEMPLATE: trip_manager
# USE FOR: travel routes, trip planning, маршруты, путешествия, поездки
# CUSTOMIZE: sections marked with # CUSTOMIZE

import asyncio
import calendar
import json
import logging
import os
import re
from datetime import datetime, date, timedelta
from pathlib import Path

import aiohttp
import aiosqlite
from aiogram import Bot, Dispatcher, F, Router
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import (
    CallbackQuery, FSInputFile, InlineKeyboardButton, InlineKeyboardMarkup,
    KeyboardButton, Message, ReplyKeyboardMarkup, ReplyKeyboardRemove,
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ── CUSTOMIZE ────────────────────────────────────────────────────────────────
BOT_DESCRIPTION = "Менеджер маршрутов и путешествий. Храню перелёты, отели, трансферы, номера броней, цены и предоплаты."
WELCOME_TEXT = (
    "🗺 <b>Менеджер маршрутов</b>\n\n"
    "Систематизирую всю информацию по поездкам:\n"
    "✈️ Перелёты · 🏨 Отели · 🚗 Трансферы\n"
    "🎯 Активности · 💰 Расходы · 📋 Задачи\n\n"
    "Выберите действие:"
)
# ── END CUSTOMIZE ─────────────────────────────────────────────────────────────

BOT_NAME = Path(__file__).stem
DATA_DIR = Path(os.getenv("DATA_DIR", "./data"))
DATA_DIR.mkdir(exist_ok=True)
DB_PATH = str(DATA_DIR / f"{BOT_NAME}_data.db")
EXCEL_PATH = str(DATA_DIR / f"{BOT_NAME}_data.xlsx")
WELCOME_IMAGE = DATA_DIR / "bot_images" / f"{BOT_NAME}.jpg"
ADMINS_FILE = DATA_DIR / f"admins_{BOT_NAME}.json"
TELEGRAPH_API = "https://api.telegra.ph"

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
router = Router()

ITEM_TYPES = {
    "flight":   "✈️ Перелёт",
    "hotel":    "🏨 Отель",
    "transfer": "🚗 Трансфер",
    "activity": "🎯 Активность",
    "expense":  "💰 Расход",
    "task":     "📋 Задача",
    "other":    "📌 Другое",
}

MONTHS_RU = {1:"янв",2:"фев",3:"мар",4:"апр",5:"май",6:"июн",
             7:"июл",8:"авг",9:"сен",10:"окт",11:"ноя",12:"дек"}

EDIT_FIELD_LABELS = {
    "title":       ("✏️ Название",        "Введите новое название:"),
    "destination": ("📍 Место",           "Введите место/город:"),
    "date_start":  ("📅 Дата начала",     "Введите дату (25.07.2025):"),
    "time_start":  ("⏰ Время",           "Введите время (14:30):"),
    "date_end":    ("📅 Дата окончания",  "Введите дату окончания (28.07.2025):"),
    "price":       ("💰 Стоимость",       "Введите сумму (цифры):"),
    "prepayment":  ("💳 Предоплата",      "Введите предоплату (цифры):"),
    "confirm_num": ("🔖 Подтверждение",   "Введите номер брони/подтверждения:"),
    "notes":       ("📝 Заметки",         "Введите заметки:"),
}


# ── admin helpers ─────────────────────────────────────────────────────────────

def _load_admins() -> set:
    try:
        return set(json.loads(ADMINS_FILE.read_text()).get("ids", []))
    except Exception:
        return set()

def _save_admins(ids: set) -> None:
    ADMINS_FILE.write_text(json.dumps({"ids": list(ids)}, ensure_ascii=False))


# ── db ────────────────────────────────────────────────────────────────────────

async def init_db():
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("""
            CREATE TABLE IF NOT EXISTS trips (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                name       TEXT NOT NULL,
                created_at TEXT DEFAULT (datetime('now','localtime'))
            )
        """)
        await db.execute("""
            CREATE TABLE IF NOT EXISTS user_prefs (
                user_id        TEXT PRIMARY KEY,
                active_trip_id INTEGER,
                digest_time    TEXT,
                digest_enabled INTEGER DEFAULT 0
            )
        """)
        await db.execute("""
            CREATE TABLE IF NOT EXISTS items (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                trip_id     INTEGER NOT NULL REFERENCES trips(id),
                item_type   TEXT DEFAULT 'other',
                title       TEXT NOT NULL,
                destination TEXT,
                date_start  TEXT,
                time_start  TEXT,
                date_end    TEXT,
                link        TEXT,
                confirm_num TEXT,
                price       REAL,
                prepayment  REAL,
                currency    TEXT DEFAULT 'RUB',
                notes       TEXT,
                remind_at   TEXT,
                status      TEXT DEFAULT 'active',
                created_at  TEXT DEFAULT (datetime('now','localtime'))
            )
        """)
        await db.execute(
            "CREATE TABLE IF NOT EXISTS _meta (key TEXT PRIMARY KEY, value TEXT)"
        )
        # migrations for existing DBs
        for table, col, definition in [
            ("items",      "time_start",    "TEXT"),
            ("user_prefs", "digest_time",   "TEXT"),
            ("user_prefs", "digest_enabled","INTEGER DEFAULT 0"),
        ]:
            try:
                await db.execute(f"ALTER TABLE {table} ADD COLUMN {col} {definition}")
            except Exception:
                pass
        await db.commit()


# ── trip helpers ──────────────────────────────────────────────────────────────

async def _all_trips() -> list:
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        rows = await (await db.execute("SELECT * FROM trips ORDER BY id DESC")).fetchall()
        return [dict(r) for r in rows]

async def _get_active_trip(user_id: str) -> dict | None:
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        row = await (await db.execute(
            "SELECT t.* FROM user_prefs p JOIN trips t ON p.active_trip_id=t.id WHERE p.user_id=?",
            (user_id,)
        )).fetchone()
        if row:
            return dict(row)
        row = await (await db.execute("SELECT * FROM trips ORDER BY id DESC LIMIT 1")).fetchone()
        return dict(row) if row else None

async def _set_active_trip(user_id: str, trip_id: int):
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute(
            "INSERT INTO user_prefs (user_id, active_trip_id) VALUES (?,?) "
            "ON CONFLICT(user_id) DO UPDATE SET active_trip_id=excluded.active_trip_id",
            (user_id, trip_id)
        )
        await db.commit()

async def _trip_items(trip_id: int, status=None) -> list:
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        if status:
            cur = await db.execute(
                "SELECT * FROM items WHERE trip_id=? AND status=? ORDER BY date_start,time_start,id",
                (trip_id, status)
            )
        else:
            cur = await db.execute(
                "SELECT * FROM items WHERE trip_id=? ORDER BY date_start,time_start,id", (trip_id,)
            )
        return [dict(r) for r in await cur.fetchall()]


# ── validation ────────────────────────────────────────────────────────────────

def _parse_date(text: str) -> str | None:
    text = text.strip()
    m = re.fullmatch(r"(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{4})", text)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
    else:
        m = re.fullmatch(r"(\d{4})[.\-](\d{1,2})[.\-](\d{1,2})", text)
        if m:
            y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        else:
            return None
    if not (2000 <= y <= 2100): return None
    if not (1 <= mo <= 12): return None
    if not (1 <= d <= calendar.monthrange(y, mo)[1]): return None
    return f"{y:04d}-{mo:02d}-{d:02d}"

def _parse_time(text: str) -> str | None:
    text = text.strip()
    m = re.fullmatch(r"(\d{1,2})[:\.](\d{2})", text)
    if not m: return None
    h, mi = int(m.group(1)), int(m.group(2))
    if not (0 <= h <= 23 and 0 <= mi <= 59): return None
    return f"{h:02d}:{mi:02d}"

def _fmt_date_ru(iso_date: str) -> str:
    try:
        dt = datetime.strptime(iso_date, "%Y-%m-%d")
        return f"{dt.day} {MONTHS_RU[dt.month]} {dt.year}"
    except Exception:
        return iso_date


# ── keyboards ─────────────────────────────────────────────────────────────────

def kb_main() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(keyboard=[
        [KeyboardButton(text="➕ Добавить"),    KeyboardButton(text="📋 Маршрут")],
        [KeyboardButton(text="🔍 Поиск"),       KeyboardButton(text="📊 Итоги")],
        [KeyboardButton(text="🗂 Путешествия"), KeyboardButton(text="📥 Excel")],
    ], resize_keyboard=True)

def kb_types() -> InlineKeyboardMarkup:
    rows = [[InlineKeyboardButton(text=v, callback_data=f"itype:{k}")] for k, v in ITEM_TYPES.items()]
    rows.append([InlineKeyboardButton(text="❌ Отмена", callback_data="icancel")])
    return InlineKeyboardMarkup(inline_keyboard=rows)

def kb_skip(step: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="⏭ Пропустить", callback_data=f"iskip:{step}"),
        InlineKeyboardButton(text="❌ Отмена",     callback_data="icancel"),
    ]])

def kb_item(item_id: int, status: str = "active") -> InlineKeyboardMarkup:
    if status == "done":
        action_btn = InlineKeyboardButton(text="↩️ Вернуть", callback_data=f"iundone:{item_id}")
    else:
        action_btn = InlineKeyboardButton(text="✅ Выполнено", callback_data=f"idone:{item_id}")
    return InlineKeyboardMarkup(inline_keyboard=[
        [action_btn, InlineKeyboardButton(text="🗑 Удалить", callback_data=f"idel:{item_id}")],
        [InlineKeyboardButton(text="✏️ Изменить", callback_data=f"iedit:{item_id}")],
        [InlineKeyboardButton(text="◀️ Назад",    callback_data="iback")],
    ])

def kb_confirm_del(item_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="🗑 Да, удалить", callback_data=f"idel_ok:{item_id}"),
        InlineKeyboardButton(text="◀️ Отмена",      callback_data=f"iview:{item_id}"),
    ]])

def kb_trips(trips: list) -> InlineKeyboardMarkup:
    rows = [[InlineKeyboardButton(text=f"🗺 {t['name']}", callback_data=f"trip_sel:{t['id']}")] for t in trips]
    rows.append([InlineKeyboardButton(text="➕ Новое путешествие", callback_data="trip_new")])
    rows.append([InlineKeyboardButton(text="❌ Закрыть", callback_data="trip_panel_close")])
    return InlineKeyboardMarkup(inline_keyboard=rows)

def kb_edit_fields(item_id: int) -> InlineKeyboardMarkup:
    fields = list(EDIT_FIELD_LABELS.items())
    rows = []
    for i in range(0, len(fields), 2):
        row = [InlineKeyboardButton(text=fields[i][1][0], callback_data=f"iedit_f:{item_id}:{fields[i][0]}")]
        if i + 1 < len(fields):
            row.append(InlineKeyboardButton(text=fields[i+1][1][0], callback_data=f"iedit_f:{item_id}:{fields[i+1][0]}"))
        rows.append(row)
    rows.append([InlineKeyboardButton(text="◀️ Назад", callback_data=f"iview:{item_id}")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


# ── helpers ───────────────────────────────────────────────────────────────────

def _fmt_price(price, prepay, currency="RUB") -> str:
    sym = {"RUB": "₽", "USD": "$", "EUR": "€", "TRY": "₺"}.get(currency, currency)
    parts = []
    if price: parts.append(f"{price:,.0f} {sym}")
    if prepay: parts.append(f"предоплата {prepay:,.0f} {sym}")
    return " · ".join(parts) if parts else "—"

def _fmt_item(r: dict) -> str:
    icon = ITEM_TYPES.get(r["item_type"], "📌").split()[0]
    lines = [f"{icon} <b>{r['title']}</b>"]
    if r.get("destination"): lines.append(f"📍 {r['destination']}")
    if r.get("date_start"):
        d_line = f"📅 {r['date_start']}"
        if r.get("time_start"): d_line += f" в {r['time_start']}"
        if r.get("date_end") and r["date_end"] != r["date_start"]:
            d_line += f" – {r['date_end']}"
        lines.append(d_line)
    p = _fmt_price(r.get("price"), r.get("prepayment"), r.get("currency", "RUB"))
    if p != "—": lines.append(f"💰 {p}")
    if r.get("confirm_num"): lines.append(f"🔖 <code>{r['confirm_num']}</code>")
    if r.get("link"):        lines.append(f"🔗 {r['link']}")
    if r.get("notes"):       lines.append(f"📝 {r['notes']}")
    if r.get("status") == "done": lines.append("✅ Выполнено")
    return "\n".join(lines)

async def _wiz_next(bot: Bot, data: dict, text: str, markup=None):
    try:
        await bot.edit_message_text(
            text, chat_id=data["wiz_chat"], message_id=data["wiz_id"],
            parse_mode="HTML", reply_markup=markup
        )
    except Exception as e:
        logger.warning(f"wiz_next failed: {e}")

async def _wiz_error(bot: Bot, data: dict, error: str, prompt: str, markup=None):
    await _wiz_next(bot, data, f"❌ {error}\n\n{prompt}", markup)


# ── FSM ───────────────────────────────────────────────────────────────────────

class Add(StatesGroup):
    type = State(); title = State(); destination = State()
    date_start = State(); time_start = State(); date_end = State()
    link = State(); confirm_num = State(); price = State()
    prepayment = State(); notes = State(); remind = State()

class Search(StatesGroup):
    q = State()

class TripCreate(StatesGroup):
    name = State()

class EditItem(StatesGroup):
    value = State()

_STEPS = [
    ("destination", Add.destination, "📍 <b>Куда / место</b> (город, адрес):"),
    ("date_start",  Add.date_start,  "📅 <b>Дата начала</b> (например: 25.07.2025):"),
    ("time_start",  Add.time_start,  "⏰ <b>Время начала</b> (например: 14:30):"),
    ("date_end",    Add.date_end,    "📅 <b>Дата окончания</b> (например: 28.07.2025):"),
    ("link",        Add.link,        "🔗 <b>Ссылка на бронирование:</b>"),
    ("confirm_num", Add.confirm_num, "🔖 <b>Номер подтверждения / брони:</b>"),
    ("price",       Add.price,       "💰 <b>Стоимость</b> (только цифры, например 15000):"),
    ("prepayment",  Add.prepayment,  "💳 <b>Предоплата</b> (цифрами, если была):"),
    ("notes",       Add.notes,       "📝 <b>Заметки / пометки:</b>"),
    ("remind",      Add.remind,      "⏰ <b>Напомнить за</b> (например: 1д / 3ч / 30м):"),
]
_STEP_MAP = {s[0]: i for i, s in enumerate(_STEPS)}


# ── /start ────────────────────────────────────────────────────────────────────

@router.message(Command("start"))
async def cmd_start(message: Message):
    admins = _load_admins()
    if not admins:
        _save_admins({str(message.from_user.id)})
    if WELCOME_IMAGE.exists():
        await message.answer_photo(FSInputFile(str(WELCOME_IMAGE)),
                                   caption=WELCOME_TEXT, parse_mode="HTML", reply_markup=kb_main())
    else:
        await message.answer(WELCOME_TEXT, parse_mode="HTML", reply_markup=kb_main())
    if str(message.from_user.id) in _load_admins():
        await message.answer(
            "🔧 <b>Команды:</b> /excel · /publish · /weblink\n"
            "/digest 08:00 — ежедневный дайджест\n"
            "/admins · /addadmin · /removeadmin",
            parse_mode="HTML"
        )


# ── TRIPS PANEL ───────────────────────────────────────────────────────────────

@router.message(F.text == "🗂 Путешествия")
async def trips_panel(msg: Message):
    trips = await _all_trips()
    if not trips:
        await msg.answer(
            "У вас ещё нет путешествий. Создайте первое!",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="➕ Новое путешествие", callback_data="trip_new")],
            ])
        )
        return
    active = await _get_active_trip(str(msg.from_user.id))
    header = f"🗺 <b>Текущее:</b> {active['name']}\n\n" if active else ""
    await msg.answer(header + "Выберите путешествие:", parse_mode="HTML", reply_markup=kb_trips(trips))

@router.callback_query(F.data == "trip_list")
async def cb_trip_list(cb: CallbackQuery):
    await cb.answer()
    trips = await _all_trips()
    active = await _get_active_trip(str(cb.from_user.id))
    header = f"🗺 <b>Текущее:</b> {active['name']}\n\n" if active else ""
    await cb.message.edit_text(header + "Выберите путешествие:", parse_mode="HTML", reply_markup=kb_trips(trips))

@router.callback_query(F.data.startswith("trip_sel:"))
async def cb_trip_sel(cb: CallbackQuery):
    await cb.answer()
    trip_id = int(cb.data.split(":")[1])
    async with aiosqlite.connect(DB_PATH) as db:
        row = await (await db.execute("SELECT name FROM trips WHERE id=?", (trip_id,))).fetchone()
    if not row:
        await cb.message.edit_text("Путешествие не найдено."); return
    await _set_active_trip(str(cb.from_user.id), trip_id)
    await cb.message.edit_text(f"✅ Активное путешествие: <b>{row[0]}</b>", parse_mode="HTML")
    await cb.message.answer("Что сделаем?", reply_markup=kb_main())

@router.callback_query(F.data == "trip_new")
async def cb_trip_new(cb: CallbackQuery, state: FSMContext):
    await cb.answer()
    await state.set_state(TripCreate.name)
    await cb.message.edit_text(
        "✏️ <b>Введите название путешествия</b>\n"
        "Например: Турция 2025, Командировка Москва, Отпуск в Сочи:",
        parse_mode="HTML"
    )

@router.message(TripCreate.name, F.text)
async def trip_create_name(msg: Message, state: FSMContext):
    name = msg.text.strip()
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute("INSERT INTO trips (name) VALUES (?)", (name,))
        trip_id = cur.lastrowid
        await db.commit()
    await _set_active_trip(str(msg.from_user.id), trip_id)
    await state.clear()
    await msg.answer(
        f"✅ Создано путешествие <b>{name}</b>!\nДобавляйте пункты — нажмите ➕ Добавить.",
        parse_mode="HTML", reply_markup=kb_main()
    )

@router.callback_query(F.data.startswith("trip_del:"))
async def cb_trip_del(cb: CallbackQuery):
    await cb.answer()
    trip_id = int(cb.data.split(":")[1])
    async with aiosqlite.connect(DB_PATH) as db:
        cnt = (await (await db.execute("SELECT COUNT(*) FROM items WHERE trip_id=?", (trip_id,))).fetchone())[0]
        row = await (await db.execute("SELECT name FROM trips WHERE id=?", (trip_id,))).fetchone()
    if not row:
        await cb.message.edit_text("Путешествие не найдено."); return
    await cb.message.edit_text(
        f"🗑 Удалить <b>{row[0]}</b>? Вместе с ним удалится <b>{cnt} пунктов</b>.",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="🗑 Да, удалить", callback_data=f"trip_del_ok:{trip_id}"),
            InlineKeyboardButton(text="◀️ Отмена",      callback_data="trip_list"),
        ]])
    )

@router.callback_query(F.data.startswith("trip_del_ok:"))
async def cb_trip_del_ok(cb: CallbackQuery):
    await cb.answer()
    trip_id = int(cb.data.split(":")[1])
    async with aiosqlite.connect(DB_PATH) as db:
        row = await (await db.execute("SELECT name FROM trips WHERE id=?", (trip_id,))).fetchone()
        await db.execute("DELETE FROM items WHERE trip_id=?", (trip_id,))
        await db.execute("DELETE FROM trips WHERE id=?", (trip_id,))
        await db.execute("UPDATE user_prefs SET active_trip_id=NULL WHERE active_trip_id=?", (trip_id,))
        await db.commit()
    await cb.message.edit_text(f"🗑 Путешествие <b>{row[0] if row else '—'}</b> удалено.", parse_mode="HTML")
    await cb.message.answer("Что сделаем?", reply_markup=kb_main())

@router.callback_query(F.data == "trip_panel_close")
async def cb_trip_close(cb: CallbackQuery):
    await cb.answer()
    await cb.message.delete()


# ── ADD FLOW ──────────────────────────────────────────────────────────────────

@router.message(F.text == "➕ Добавить")
async def add_start(message: Message, state: FSMContext):
    trip = await _get_active_trip(str(message.from_user.id))
    if not trip:
        await message.answer("Сначала создайте путешествие — 🗂 Путешествия.", reply_markup=kb_main())
        return
    await state.set_state(Add.type)
    wiz = await message.answer(
        f"🗺 <b>{trip['name']}</b>\n\nВыберите тип:",
        parse_mode="HTML", reply_markup=kb_types()
    )
    await state.update_data(trip_id=trip["id"], wiz_id=wiz.message_id, wiz_chat=wiz.chat.id)

@router.callback_query(F.data.startswith("itype:"))
async def cb_type(cb: CallbackQuery, state: FSMContext):
    await cb.answer()
    itype = cb.data.split(":")[1]
    await state.update_data(item_type=itype)
    await state.set_state(Add.title)
    await cb.message.edit_text(
        f"{ITEM_TYPES[itype]}\n\n✏️ <b>Название</b> (например: «Рейс SU1234», «Hilton Istanbul»):",
        parse_mode="HTML"
    )

@router.message(Add.title, F.text)
async def add_title(msg: Message, state: FSMContext, bot: Bot):
    await state.update_data(title=msg.text.strip())
    await state.set_state(Add.destination)
    data = await state.get_data()
    await _wiz_next(bot, data, "📍 <b>Куда / место</b> (город, адрес):", kb_skip("destination"))

@router.message(Add.destination, F.text, ~F.text.startswith("/"))
async def add_destination(msg: Message, state: FSMContext, bot: Bot):
    await state.update_data(destination=msg.text.strip())
    await state.set_state(Add.date_start)
    data = await state.get_data()
    await _wiz_next(bot, data, "📅 <b>Дата начала</b> (например: 25.07.2025):", kb_skip("date_start"))

@router.message(Add.date_start, F.text, ~F.text.startswith("/"))
async def add_date_start(msg: Message, state: FSMContext, bot: Bot):
    parsed = _parse_date(msg.text)
    data = await state.get_data()
    if parsed is None:
        await _wiz_error(bot, data, "Неверный формат даты.",
                         "📅 <b>Дата начала</b> (например: 25.07.2025):", kb_skip("date_start"))
        return
    await state.update_data(date_start=parsed)
    await state.set_state(Add.time_start)
    await _wiz_next(bot, data, "⏰ <b>Время начала</b> (например: 14:30):", kb_skip("time_start"))

@router.message(Add.time_start, F.text, ~F.text.startswith("/"))
async def add_time_start(msg: Message, state: FSMContext, bot: Bot):
    parsed = _parse_time(msg.text)
    data = await state.get_data()
    if parsed is None:
        await _wiz_error(bot, data, "Неверный формат. Введите время как 14:30.",
                         "⏰ <b>Время начала</b> (например: 14:30):", kb_skip("time_start"))
        return
    await state.update_data(time_start=parsed)
    await state.set_state(Add.date_end)
    await _wiz_next(bot, data, "📅 <b>Дата окончания</b> (например: 28.07.2025):", kb_skip("date_end"))

@router.message(Add.date_end, F.text, ~F.text.startswith("/"))
async def add_date_end(msg: Message, state: FSMContext, bot: Bot):
    parsed = _parse_date(msg.text)
    data = await state.get_data()
    if parsed is None:
        await _wiz_error(bot, data, "Неверный формат даты.",
                         "📅 <b>Дата окончания</b> (например: 28.07.2025):", kb_skip("date_end"))
        return
    await state.update_data(date_end=parsed)
    await state.set_state(Add.link)
    await _wiz_next(bot, data, "🔗 <b>Ссылка на бронирование:</b>", kb_skip("link"))

@router.message(Add.link, F.text, ~F.text.startswith("/"))
async def add_link(msg: Message, state: FSMContext, bot: Bot):
    await state.update_data(link=msg.text.strip())
    await state.set_state(Add.confirm_num)
    data = await state.get_data()
    await _wiz_next(bot, data, "🔖 <b>Номер подтверждения / брони:</b>", kb_skip("confirm_num"))

@router.message(Add.confirm_num, F.text, ~F.text.startswith("/"))
async def add_confirm(msg: Message, state: FSMContext, bot: Bot):
    await state.update_data(confirm_num=msg.text.strip())
    await state.set_state(Add.price)
    data = await state.get_data()
    await _wiz_next(bot, data, "💰 <b>Стоимость</b> (только цифры, например 15000):", kb_skip("price"))

@router.message(Add.price, F.text, ~F.text.startswith("/"))
async def add_price(msg: Message, state: FSMContext, bot: Bot):
    cleaned = re.sub(r"[^\d.]", "", msg.text)
    data = await state.get_data()
    try:
        await state.update_data(price=float(cleaned))
    except ValueError:
        await _wiz_error(bot, data, "Введите число.",
                         "💰 <b>Стоимость</b> (только цифры, например 15000):", kb_skip("price"))
        return
    await state.set_state(Add.prepayment)
    await _wiz_next(bot, data, "💳 <b>Предоплата</b> (цифрами, если была):", kb_skip("prepayment"))

@router.message(Add.prepayment, F.text, ~F.text.startswith("/"))
async def add_prepay(msg: Message, state: FSMContext, bot: Bot):
    cleaned = re.sub(r"[^\d.]", "", msg.text)
    data = await state.get_data()
    try:
        await state.update_data(prepayment=float(cleaned))
    except ValueError:
        await _wiz_error(bot, data, "Введите число.",
                         "💳 <b>Предоплата</b> (цифрами, если была):", kb_skip("prepayment"))
        return
    await state.set_state(Add.notes)
    await _wiz_next(bot, data, "📝 <b>Заметки / пометки:</b>", kb_skip("notes"))

@router.message(Add.notes, F.text, ~F.text.startswith("/"))
async def add_notes(msg: Message, state: FSMContext, bot: Bot):
    await state.update_data(notes=msg.text.strip())
    await state.set_state(Add.remind)
    data = await state.get_data()
    await _wiz_next(bot, data, "⏰ <b>Напомнить за</b> (например: 1д / 3ч / 30м):", kb_skip("remind"))

@router.message(Add.remind, F.text, ~F.text.startswith("/"))
async def add_remind(msg: Message, state: FSMContext, bot: Bot):
    await state.update_data(remind_raw=msg.text.strip())
    await _finalize(msg, state, bot)


async def _finalize(msg: Message, state: FSMContext, bot: Bot):
    data = await state.get_data()
    remind_raw = data.pop("remind_raw", None)
    remind_at = None
    date_start = data.get("date_start")
    if remind_raw and date_start:
        try:
            d = datetime.strptime(date_start, "%Y-%m-%d")
            m = re.match(r"(\d+)([дdhHчcm])", remind_raw.lower())
            if m:
                n, u = int(m.group(1)), m.group(2)
                delta = (timedelta(days=n) if u in "дd"
                         else timedelta(hours=n) if u in "чhc"
                         else timedelta(minutes=n))
                remind_at = (d - delta).isoformat()
        except Exception:
            pass

    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute(
            "INSERT INTO items "
            "(trip_id,item_type,title,destination,date_start,time_start,date_end,"
            "link,confirm_num,price,prepayment,notes,remind_at) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (data.get("trip_id"), data.get("item_type", "other"), data.get("title", ""),
             data.get("destination"), data.get("date_start"), data.get("time_start"),
             data.get("date_end"), data.get("link"), data.get("confirm_num"),
             data.get("price"), data.get("prepayment"), data.get("notes"), remind_at)
        )
        await db.commit()

    icon = ITEM_TYPES.get(data.get("item_type", "other"), "📌").split()[0]
    lines = [f"✅ <b>Сохранено!</b>\n", f"{icon} {data.get('title', '')}"]
    if data.get("destination"): lines.append(f"📍 {data['destination']}")
    if data.get("date_start"):
        d_line = f"📅 {data['date_start']}"
        if data.get("time_start"): d_line += f" в {data['time_start']}"
        lines.append(d_line)
    if data.get("price"): lines.append(f"💰 {data['price']:,.0f} ₽")
    success = "\n".join(lines)

    wiz_id = data.get("wiz_id")
    wiz_chat = data.get("wiz_chat")
    await state.clear()

    if wiz_id and wiz_chat:
        try:
            await bot.edit_message_text(success, chat_id=wiz_chat, message_id=wiz_id, parse_mode="HTML")
            await bot.send_message(wiz_chat, "Что сделаем?", reply_markup=kb_main())
            if remind_at:
                asyncio.create_task(_remind(bot, wiz_chat, data.get("title", ""), remind_at))
            return
        except Exception:
            pass
    await msg.answer(success, parse_mode="HTML", reply_markup=kb_main())
    if remind_at:
        asyncio.create_task(_remind(bot, msg.chat.id, data.get("title", ""), remind_at))


async def _remind(bot: Bot, chat_id: int, title: str, remind_at: str):
    try:
        delay = (datetime.fromisoformat(remind_at) - datetime.now()).total_seconds()
        if delay > 0:
            await asyncio.sleep(delay)
        await bot.send_message(chat_id, f"⏰ <b>Напоминание:</b> {title}", parse_mode="HTML")
    except Exception as e:
        logger.error(f"Reminder error: {e}")


# ── SKIP ──────────────────────────────────────────────────────────────────────

@router.callback_query(F.data.startswith("iskip:"))
async def cb_skip(cb: CallbackQuery, state: FSMContext, bot: Bot):
    await cb.answer()
    step = cb.data.split(":")[1]
    if step == "remind":
        await _finalize(cb.message, state, bot)
        return
    idx = _STEP_MAP.get(step, -1)
    if idx + 1 < len(_STEPS):
        next_step, next_state, prompt = _STEPS[idx + 1]
        await state.set_state(next_state)
        data = await state.get_data()
        await _wiz_next(bot, data, prompt, kb_skip(next_step))

@router.callback_query(F.data == "icancel")
async def cb_icancel(cb: CallbackQuery, state: FSMContext):
    await cb.answer()
    await state.clear()
    await cb.message.edit_text("Отменено.")
    await cb.message.answer("Что сделаем?", reply_markup=kb_main())


# ── ITEM EDIT ─────────────────────────────────────────────────────────────────

@router.callback_query(F.data.startswith("iedit:"))
async def cb_iedit(cb: CallbackQuery):
    await cb.answer()
    item_id = int(cb.data.split(":")[1])
    await cb.message.edit_text(
        "✏️ <b>Что изменить?</b>",
        parse_mode="HTML",
        reply_markup=kb_edit_fields(item_id)
    )

@router.callback_query(F.data.startswith("iedit_f:"))
async def cb_iedit_field(cb: CallbackQuery, state: FSMContext):
    await cb.answer()
    parts = cb.data.split(":", 2)
    item_id, field = int(parts[1]), parts[2]
    _, prompt = EDIT_FIELD_LABELS[field]
    await state.set_state(EditItem.value)
    await state.update_data(
        edit_item_id=item_id, edit_field=field,
        edit_msg_id=cb.message.message_id, edit_chat_id=cb.message.chat.id
    )
    await cb.message.edit_text(
        f"✏️ {prompt}",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="❌ Отмена", callback_data=f"iedit:{item_id}")
        ]])
    )

@router.message(EditItem.value, F.text)
async def edit_item_value(msg: Message, state: FSMContext, bot: Bot):
    data = await state.get_data()
    field = data["edit_field"]
    item_id = data["edit_item_id"]
    value = msg.text.strip()

    if field in ("date_start", "date_end"):
        value = _parse_date(value)
        if not value:
            try:
                await bot.edit_message_text(
                    "❌ Неверный формат даты. Введите: 25.07.2025\n\nПовторите:",
                    chat_id=data["edit_chat_id"], message_id=data["edit_msg_id"]
                )
            except Exception:
                pass
            return
    elif field == "time_start":
        value = _parse_time(value)
        if not value:
            try:
                await bot.edit_message_text(
                    "❌ Неверный формат. Введите: 14:30\n\nПовторите:",
                    chat_id=data["edit_chat_id"], message_id=data["edit_msg_id"]
                )
            except Exception:
                pass
            return
    elif field in ("price", "prepayment"):
        cleaned = re.sub(r"[^\d.]", "", value)
        try:
            value = float(cleaned)
        except ValueError:
            try:
                await bot.edit_message_text(
                    "❌ Введите число.\n\nПовторите:",
                    chat_id=data["edit_chat_id"], message_id=data["edit_msg_id"]
                )
            except Exception:
                pass
            return

    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute(f"UPDATE items SET {field}=? WHERE id=?", (value, item_id))
        await db.commit()
        db.row_factory = aiosqlite.Row
        row = await (await db.execute("SELECT * FROM items WHERE id=?", (item_id,))).fetchone()

    await state.clear()
    if row:
        r = dict(row)
        try:
            await bot.edit_message_text(
                "✅ Сохранено!\n\n" + _fmt_item(r),
                chat_id=data["edit_chat_id"], message_id=data["edit_msg_id"],
                parse_mode="HTML", reply_markup=kb_item(item_id, r.get("status", "active"))
            )
            return
        except Exception:
            pass
    await msg.answer("✅ Изменено.", reply_markup=kb_main())


# ── ROUTE VIEW ────────────────────────────────────────────────────────────────

@router.message(F.text == "📋 Маршрут")
async def show_route(msg: Message):
    trip = await _get_active_trip(str(msg.from_user.id))
    if not trip:
        await msg.answer("Сначала создайте путешествие — 🗂 Путешествия.", reply_markup=kb_main())
        return
    items = await _trip_items(trip["id"], "active")
    if not items:
        await msg.answer(
            f"🗺 <b>{trip['name']}</b>\n\nМаршрут пуст. Нажмите ➕ Добавить.",
            parse_mode="HTML", reply_markup=kb_main()
        )
        return

    by_date: dict = {}
    no_date = []
    for item in items:
        d = item.get("date_start") or ""
        if d:
            by_date.setdefault(d, []).append(item)
        else:
            no_date.append(item)

    lines = [f"🗺 <b>{trip['name']}</b>\n"]
    for d in sorted(by_date):
        try:
            dt = datetime.strptime(d, "%Y-%m-%d")
            lines.append(f"\n<b>📅 {dt.strftime('%d.%m.%Y')}</b>")
        except Exception:
            lines.append(f"\n<b>📅 {d}</b>")
        for i in by_date[d]:
            icon = ITEM_TYPES.get(i["item_type"], "📌").split()[0]
            time_str = f"{i['time_start']} " if i.get("time_start") else ""
            line = f"  {time_str}{icon} {i['title']}"
            p = _fmt_price(i.get("price"), i.get("prepayment"))
            if p != "—": line += f" · {p}"
            if i.get("confirm_num"): line += f"\n     🔖 <code>{i['confirm_num']}</code>"
            lines.append(line)
    if no_date:
        lines.append("\n<b>📌 Без даты</b>")
        for i in no_date:
            lines.append(f"  {ITEM_TYPES.get(i['item_type'], '📌').split()[0]} {i['title']}")

    total_price = sum(i.get("price") or 0 for i in items)
    total_pre   = sum(i.get("prepayment") or 0 for i in items)
    lines.append(f"\n💼 Итого: {len(items)} пунктов")
    if total_price:
        lines.append(f"💰 Сумма: {total_price:,.0f} ₽  |  Предоплачено: {total_pre:,.0f} ₽")
        lines.append(f"📌 Осталось оплатить: {total_price - total_pre:,.0f} ₽")

    btn_rows = []
    for i in items[:15]:
        btn_rows.append([InlineKeyboardButton(
            text=f"{ITEM_TYPES.get(i['item_type'], '📌').split()[0]} {i['title'][:28]}",
            callback_data=f"iview:{i['id']}"
        )])
    btn_rows.append([
        InlineKeyboardButton(text="📅 По дням",    callback_data=f"route_days:{trip['id']}"),
        InlineKeyboardButton(text="🗓 Эта неделя", callback_data=f"route_week:{trip['id']}"),
    ])

    text = "\n".join(lines)
    if len(text) > 4000: text = text[:3900] + "\n…"
    await msg.answer(text, parse_mode="HTML", reply_markup=InlineKeyboardMarkup(inline_keyboard=btn_rows))


@router.callback_query(F.data.startswith("route_days:"))
async def cb_route_days(cb: CallbackQuery):
    await cb.answer()
    trip_id = int(cb.data.split(":")[1])
    items = await _trip_items(trip_id, "active")
    dates = sorted(set(i["date_start"] for i in items if i.get("date_start")))
    if not dates:
        await cb.message.answer("Нет пунктов с датами.")
        return

    today_iso    = date.today().isoformat()
    tomorrow_iso = (date.today() + timedelta(days=1)).isoformat()
    rows = []
    for d in dates:
        cnt = sum(1 for i in items if i.get("date_start") == d)
        label = _fmt_date_ru(d)
        if d == today_iso: label = f"Сегодня, {label}"
        elif d == tomorrow_iso: label = f"Завтра, {label}"
        rows.append([InlineKeyboardButton(
            text=f"📅 {label} ({cnt})",
            callback_data=f"route_day:{trip_id}:{d}"
        )])
    rows.append([InlineKeyboardButton(text="❌ Закрыть", callback_data="route_close")])

    await cb.message.answer(
        "📅 <b>Выберите день:</b>", parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=rows)
    )

@router.callback_query(F.data.startswith("route_day:"))
async def cb_route_day(cb: CallbackQuery):
    await cb.answer()
    parts = cb.data.split(":", 2)
    trip_id, day = int(parts[1]), parts[2]
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        items = [dict(r) for r in await (await db.execute(
            "SELECT * FROM items WHERE trip_id=? AND date_start=? ORDER BY time_start,id",
            (trip_id, day)
        )).fetchall()]
        trip_row = await (await db.execute("SELECT name FROM trips WHERE id=?", (trip_id,))).fetchone()
    trip_name = trip_row[0] if trip_row else ""

    today_iso    = date.today().isoformat()
    tomorrow_iso = (date.today() + timedelta(days=1)).isoformat()
    date_label = _fmt_date_ru(day)
    if day == today_iso: date_label = "Сегодня, " + date_label
    elif day == tomorrow_iso: date_label = "Завтра, " + date_label

    if not items:
        text = f"📅 <b>{date_label}</b>\n\nПунктов нет."
    else:
        lines = [f"📅 <b>{date_label}</b>  —  {trip_name}\n"]
        for item in items:
            icon = ITEM_TYPES.get(item["item_type"], "📌").split()[0]
            time_str = f"⏰ {item['time_start']}  " if item.get("time_start") else ""
            lines.append(f"{time_str}{icon} <b>{item['title']}</b>")
            if item.get("destination"): lines.append(f"   📍 {item['destination']}")
            p = _fmt_price(item.get("price"), item.get("prepayment"))
            if p != "—": lines.append(f"   💰 {p}")
            if item.get("confirm_num"): lines.append(f"   🔖 {item['confirm_num']}")
            if item.get("notes"): lines.append(f"   📝 {item['notes']}")
            lines.append("")
        text = "\n".join(lines).rstrip()

    btn_rows = [[InlineKeyboardButton(
        text=f"{ITEM_TYPES.get(i['item_type'],'📌').split()[0]} {i['title'][:28]}",
        callback_data=f"iview:{i['id']}"
    )] for i in items]
    btn_rows.append([
        InlineKeyboardButton(text="◀️ К выбору дня", callback_data=f"route_days:{trip_id}"),
        InlineKeyboardButton(text="❌ Закрыть",       callback_data="route_close"),
    ])
    await cb.message.edit_text(
        text, parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=btn_rows)
    )

@router.callback_query(F.data.startswith("route_week:"))
async def cb_route_week(cb: CallbackQuery):
    await cb.answer()
    trip_id = int(cb.data.split(":")[1])
    today = date.today()
    week_end = today + timedelta(days=7)
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        items = [dict(r) for r in await (await db.execute(
            "SELECT * FROM items WHERE trip_id=? AND date_start>=? AND date_start<=? "
            "ORDER BY date_start,time_start,id",
            (trip_id, today.isoformat(), week_end.isoformat())
        )).fetchall()]
        trip_row = await (await db.execute("SELECT name FROM trips WHERE id=?", (trip_id,))).fetchone()
    trip_name = trip_row[0] if trip_row else ""

    if not items:
        await cb.message.answer(
            f"🗓 <b>Ближайшая неделя — {trip_name}</b>\n\nНет пунктов на ближайшие 7 дней.",
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
                InlineKeyboardButton(text="❌ Закрыть", callback_data="route_close")
            ]])
        )
        return

    by_date: dict = {}
    for item in items:
        by_date.setdefault(item["date_start"], []).append(item)

    days_map = {0:"Пн",1:"Вт",2:"Ср",3:"Чт",4:"Пт",5:"Сб",6:"Вс"}
    lines = [f"🗓 <b>Ближайшая неделя — {trip_name}</b>\n"]
    for d in sorted(by_date):
        try:
            dt_date = datetime.strptime(d, "%Y-%m-%d").date()
            if dt_date == today:
                label = f"Сегодня, {dt_date.day} {MONTHS_RU[dt_date.month]}"
            elif dt_date == today + timedelta(days=1):
                label = f"Завтра, {dt_date.day} {MONTHS_RU[dt_date.month]}"
            else:
                label = f"{days_map[dt_date.weekday()]}, {dt_date.day} {MONTHS_RU[dt_date.month]}"
        except Exception:
            label = d
        lines.append(f"\n<b>📅 {label}</b>")
        for item in by_date[d]:
            icon = ITEM_TYPES.get(item["item_type"], "📌").split()[0]
            time_str = f"{item['time_start']} " if item.get("time_start") else ""
            line = f"  {time_str}{icon} {item['title']}"
            if item.get("destination"): line += f" → {item['destination']}"
            lines.append(line)

    text = "\n".join(lines)
    if len(text) > 4000: text = text[:3900] + "\n…"
    await cb.message.answer(
        text, parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="📅 По дням",  callback_data=f"route_days:{trip_id}"),
            InlineKeyboardButton(text="❌ Закрыть", callback_data="route_close"),
        ]])
    )

@router.callback_query(F.data == "route_close")
async def cb_route_close(cb: CallbackQuery):
    await cb.answer()
    await cb.message.delete()


@router.callback_query(F.data.startswith("iview:"))
async def cb_view(cb: CallbackQuery):
    await cb.answer()
    item_id = int(cb.data.split(":")[1])
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        row = await (await db.execute("SELECT * FROM items WHERE id=?", (item_id,))).fetchone()
    if not row:
        await cb.message.answer("Не найдено.")
        return
    r = dict(row)
    await cb.message.answer(_fmt_item(r), parse_mode="HTML", reply_markup=kb_item(item_id, r.get("status", "active")))

@router.callback_query(F.data.startswith("idel:"))
async def cb_del(cb: CallbackQuery):
    await cb.answer()
    await cb.message.edit_reply_markup(reply_markup=kb_confirm_del(int(cb.data.split(":")[1])))

@router.callback_query(F.data.startswith("idel_ok:"))
async def cb_del_ok(cb: CallbackQuery):
    await cb.answer()
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("DELETE FROM items WHERE id=?", (int(cb.data.split(":")[1]),))
        await db.commit()
    await cb.message.edit_text("🗑 Удалено.")

@router.callback_query(F.data.startswith("idone:"))
async def cb_done(cb: CallbackQuery):
    await cb.answer()
    item_id = int(cb.data.split(":")[1])
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("UPDATE items SET status='done' WHERE id=?", (item_id,))
        await db.commit()
    await cb.message.edit_text("✅ Отмечено как выполнено!", reply_markup=kb_item(item_id, "done"))

@router.callback_query(F.data.startswith("iundone:"))
async def cb_undone(cb: CallbackQuery):
    await cb.answer()
    item_id = int(cb.data.split(":")[1])
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("UPDATE items SET status='active' WHERE id=?", (item_id,))
        await db.commit()
    await cb.message.edit_text("↩️ Возвращено в активные.", reply_markup=kb_item(item_id, "active"))

@router.callback_query(F.data == "iback")
async def cb_iback(cb: CallbackQuery):
    await cb.answer()
    await cb.message.delete()


# ── SEARCH ────────────────────────────────────────────────────────────────────

@router.message(F.text == "🔍 Поиск")
async def search_start(msg: Message, state: FSMContext):
    await state.set_state(Search.q)
    await msg.answer("🔍 Введите слово (название, город, номер брони):", reply_markup=ReplyKeyboardRemove())

@router.message(Search.q, F.text)
async def search_do(msg: Message, state: FSMContext):
    await state.clear()
    trip = await _get_active_trip(str(msg.from_user.id))
    q = f"%{msg.text.strip()}%"
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        if trip:
            rows = await (await db.execute(
                "SELECT * FROM items WHERE trip_id=? AND "
                "(title LIKE ? OR destination LIKE ? OR confirm_num LIKE ? OR notes LIKE ?)",
                (trip["id"], q, q, q, q)
            )).fetchall()
        else:
            rows = await (await db.execute(
                "SELECT * FROM items WHERE title LIKE ? OR destination LIKE ? OR confirm_num LIKE ? OR notes LIKE ?",
                (q, q, q, q)
            )).fetchall()
    if not rows:
        await msg.answer("Ничего не найдено.", reply_markup=kb_main())
        return
    lines = [f"🔍 Найдено: {len(rows)}\n"]
    for r in [dict(x) for x in rows[:10]]:
        icon = ITEM_TYPES.get(r["item_type"], "📌").split()[0]
        lines.append(f"{icon} <b>{r['title']}</b>")
        if r.get("date_start"):
            d_line = f"   📅 {r['date_start']}"
            if r.get("time_start"): d_line += f" в {r['time_start']}"
            lines.append(d_line)
        if r.get("confirm_num"): lines.append(f"   🔖 <code>{r['confirm_num']}</code>")
    await msg.answer("\n".join(lines), parse_mode="HTML", reply_markup=kb_main())


# ── STATS ─────────────────────────────────────────────────────────────────────

@router.message(F.text == "📊 Итоги")
async def show_stats(msg: Message):
    trips = await _all_trips()
    if not trips:
        await msg.answer("Нет путешествий. Нажмите 🗂 Путешествия чтобы создать.")
        return
    if len(trips) == 1:
        await _show_trip_stats(msg, trips[0])
        return
    rows = [[InlineKeyboardButton(text=f"🗺 {t['name']}", callback_data=f"stats_trip:{t['id']}")] for t in trips]
    await msg.answer("📊 Выберите путешествие для итогов:", reply_markup=InlineKeyboardMarkup(inline_keyboard=rows))

@router.callback_query(F.data.startswith("stats_trip:"))
async def cb_stats_trip(cb: CallbackQuery):
    await cb.answer()
    trip_id = int(cb.data.split(":")[1])
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        row = await (await db.execute("SELECT * FROM trips WHERE id=?", (trip_id,))).fetchone()
    if not row:
        await cb.message.edit_text("Путешествие не найдено.")
        return
    await cb.message.delete()
    await _show_trip_stats(cb.message, dict(row))

async def _show_trip_stats(msg: Message, trip: dict):
    async with aiosqlite.connect(DB_PATH) as db:
        total = (await (await db.execute(
            "SELECT COUNT(*) FROM items WHERE trip_id=?", (trip["id"],))).fetchone())[0]
        done  = (await (await db.execute(
            "SELECT COUNT(*) FROM items WHERE trip_id=? AND status='done'", (trip["id"],))).fetchone())[0]
        cost  = (await (await db.execute(
            "SELECT COALESCE(SUM(price),0) FROM items WHERE trip_id=? AND status='active'",
            (trip["id"],))).fetchone())[0]
        paid  = (await (await db.execute(
            "SELECT COALESCE(SUM(prepayment),0) FROM items WHERE trip_id=? AND status='active'",
            (trip["id"],))).fetchone())[0]
        by_type = await (await db.execute(
            "SELECT item_type, COUNT(*), COALESCE(SUM(price),0) FROM items WHERE trip_id=? GROUP BY item_type",
            (trip["id"],)
        )).fetchall()
    lines = [f"📊 <b>Итоги: {trip['name']}</b>\n",
             f"📋 Пунктов: {total} (выполнено: {done})"]
    if cost:
        lines += [f"💰 Общая стоимость: {cost:,.0f} ₽",
                  f"💳 Оплачено: {paid:,.0f} ₽",
                  f"📌 Осталось оплатить: {cost - paid:,.0f} ₽"]
    lines.append("\n<b>По типам:</b>")
    for itype, cnt, ttl in by_type:
        label = ITEM_TYPES.get(itype, itype)
        lines.append(f"  {label}: {cnt}" + (f" · {ttl:,.0f} ₽" if ttl else ""))
    await msg.answer("\n".join(lines), parse_mode="HTML", reply_markup=kb_main())


# ── DAILY DIGEST ──────────────────────────────────────────────────────────────

@router.message(Command("digest"))
async def cmd_digest(msg: Message):
    parts = msg.text.split()
    if len(parts) < 2:
        await msg.answer(
            "⏰ <b>Ежедневный дайджест</b>\n\n"
            "Каждый день присылаю план на завтра.\n\n"
            "/digest 08:00 — получать утром\n"
            "/digest 20:00 — получать вечером\n"
            "/digest off — отключить",
            parse_mode="HTML"
        )
        return
    arg = parts[1].lower()
    user_id = str(msg.from_user.id)
    async with aiosqlite.connect(DB_PATH) as db:
        if arg == "off":
            await db.execute(
                "INSERT INTO user_prefs (user_id, digest_enabled) VALUES (?,0) "
                "ON CONFLICT(user_id) DO UPDATE SET digest_enabled=0",
                (user_id,)
            )
            await db.commit()
            await msg.answer("✅ Дайджест отключён.")
            return
        t = _parse_time(arg)
        if not t:
            await msg.answer("❌ Неверный формат. Пример: /digest 08:00")
            return
        await db.execute(
            "INSERT INTO user_prefs (user_id, digest_time, digest_enabled) VALUES (?,?,1) "
            "ON CONFLICT(user_id) DO UPDATE SET digest_time=excluded.digest_time, digest_enabled=1",
            (user_id, t)
        )
        await db.commit()
    await msg.answer(
        f"✅ Дайджест настроен на <b>{t}</b>\nБуду присылать план на завтра каждый день.",
        parse_mode="HTML"
    )

async def _send_digest(bot: Bot, user_id: str):
    tomorrow = (date.today() + timedelta(days=1)).isoformat()
    trip = await _get_active_trip(user_id)
    if not trip: return
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        rows = await (await db.execute(
            "SELECT * FROM items WHERE trip_id=? AND date_start=? AND status='active' ORDER BY time_start,id",
            (trip["id"], tomorrow)
        )).fetchall()
    if not rows: return
    lines = [f"🌅 <b>План на завтра — {trip['name']}</b>\n<b>{_fmt_date_ru(tomorrow)}</b>\n"]
    for r in [dict(x) for x in rows]:
        icon = ITEM_TYPES.get(r["item_type"], "📌").split()[0]
        time_str = f"⏰ {r['time_start']}  " if r.get("time_start") else ""
        line = f"{time_str}{icon} {r['title']}"
        if r.get("destination"): line += f" — {r['destination']}"
        if r.get("price"): line += f" · {r['price']:,.0f} ₽"
        lines.append(line)
    try:
        await bot.send_message(int(user_id), "\n".join(lines), parse_mode="HTML")
    except Exception as e:
        logger.error(f"Digest send error for {user_id}: {e}")

async def _digest_loop(bot: Bot):
    sent_keys: set = set()
    while True:
        try:
            now = datetime.now()
            today_str = now.date().isoformat()
            async with aiosqlite.connect(DB_PATH) as db:
                rows = await (await db.execute(
                    "SELECT user_id, digest_time FROM user_prefs "
                    "WHERE digest_enabled=1 AND digest_time IS NOT NULL"
                )).fetchall()
            for user_id, digest_time in rows:
                try:
                    h, m = map(int, digest_time.split(":"))
                except Exception:
                    continue
                key = f"{user_id}_{today_str}_{digest_time}"
                if key in sent_keys: continue
                target = now.replace(hour=h, minute=m, second=0, microsecond=0)
                if now >= target and (now - target).total_seconds() < 120:
                    await _send_digest(bot, user_id)
                    sent_keys.add(key)
            sent_keys = {k for k in sent_keys if today_str in k}
        except Exception as e:
            logger.error(f"Digest loop error: {e}")
        await asyncio.sleep(60)


# ── EXCEL ─────────────────────────────────────────────────────────────────────

@router.message(F.text == "📥 Excel")
@router.message(Command("excel"))
async def cmd_excel(msg: Message):
    if str(msg.from_user.id) not in _load_admins():
        await msg.answer("⛔ Нет доступа"); return
    trips = await _all_trips()
    if not trips:
        await msg.answer("Данных нет."); return
    wb = Workbook(); wb.remove(wb.active)
    hdrs = ["Тип", "Название", "Куда", "Дата нач.", "Время", "Дата кон.", "Ссылка",
            "Подтверждение", "Цена", "Предоплата", "Заметки", "Статус"]
    hfill  = PatternFill("solid", fgColor="1F4E79")
    hfont  = Font(bold=True, color="FFFFFF")
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"),  bottom=Side(style="thin"))
    fills      = [PatternFill("solid", fgColor="FFFFFF"), PatternFill("solid", fgColor="DCE6F1")]
    total_fill = PatternFill("solid", fgColor="FFF2CC")
    total_font = Font(bold=True)

    for trip in trips:
        items = await _trip_items(trip["id"])
        ws = wb.create_sheet(title=trip["name"][:31])
        for c, h in enumerate(hdrs, 1):
            cell = ws.cell(1, c, h)
            cell.fill = hfill; cell.font = hfont; cell.border = border
            cell.alignment = Alignment(horizontal="center")
        total_price = 0; total_pre = 0
        for ri, item in enumerate(items, 2):
            vals = [
                ITEM_TYPES.get(item["item_type"], ""), item["title"],
                item.get("destination", ""), item.get("date_start", ""), item.get("time_start", ""),
                item.get("date_end", ""), item.get("link", ""), item.get("confirm_num", ""),
                item.get("price"), item.get("prepayment"), item.get("notes", ""), item.get("status", "")
            ]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(ri, c, v); cell.fill = fills[ri % 2]; cell.border = border
            total_price += item.get("price") or 0
            total_pre   += item.get("prepayment") or 0
        tr = len(items) + 2
        for c, v in [(1, "ИТОГО"), (9, total_price), (10, total_pre),
                     (12, f"Осталось: {total_price - total_pre:,.0f} ₽")]:
            ws.cell(tr, c, v).font = total_font
            ws.cell(tr, c).fill = total_fill
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(
                max(max(len(str(c.value or "")) for c in col) + 2, 10), 40
            )
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:L{len(items) + 1}"

    wb.save(EXCEL_PATH)
    await msg.answer_document(FSInputFile(EXCEL_PATH),
                              caption=f"📊 Таблица маршрутов ({len(trips)} путешествий)")


# ── TELEGRAPH ─────────────────────────────────────────────────────────────────

async def _tg_token() -> str:
    async with aiosqlite.connect(DB_PATH) as db:
        row = await (await db.execute("SELECT value FROM _meta WHERE key='tg_token'")).fetchone()
    if row: return row[0]
    async with aiohttp.ClientSession() as s:
        async with s.post(f"{TELEGRAPH_API}/createAccount",
                          data={"short_name": BOT_NAME[:31], "author_name": "TripBot"}) as r:
            token = (await r.json())["result"]["access_token"]
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("INSERT OR REPLACE INTO _meta VALUES ('tg_token',?)", (token,))
        await db.commit()
    return token

def _tg_nodes_items(trip_name: str, items: list) -> list:
    DAYS_FULL   = {0:"понедельник",1:"вторник",2:"среда",3:"четверг",
                   4:"пятница",5:"суббота",6:"воскресенье"}
    MONTHS_FULL = {1:"января",2:"февраля",3:"марта",4:"апреля",5:"мая",6:"июня",
                   7:"июля",8:"августа",9:"сентября",10:"октября",11:"ноября",12:"декабря"}

    total_price = sum(i.get("price") or 0 for i in items)
    total_pre   = sum(i.get("prepayment") or 0 for i in items)

    nodes: list = [{"tag": "h3", "children": [trip_name]}]
    sub = f"{len(items)} событий"
    if total_price:
        sub += f"  ·  {total_price:,.0f} ₽ всего"
    nodes.append({"tag": "p", "children": [sub]})

    by_date: dict = {}
    no_date = []
    for item in items:
        d = item.get("date_start") or ""
        (by_date.setdefault(d, []) if d else no_date).append(item)

    def _item_nodes(item: dict) -> list:
        icon = ITEM_TYPES.get(item["item_type"], "📌").split()[0]
        title = f"{icon}  {item['title']}"
        if item.get("time_start"):
            title += f"  —  {item['time_start']}"
        result = [{"tag": "p", "children": [{"tag": "b", "children": [title]}]}]
        details = []
        if item.get("destination"):
            details.append({"tag": "li", "children": [f"📍  {item['destination']}"]})
        if item.get("date_end") and item["date_end"] != item.get("date_start"):
            try:
                dt2 = datetime.strptime(item["date_end"], "%Y-%m-%d")
                end_label = f"{dt2.day} {MONTHS_FULL[dt2.month]} {dt2.year}"
            except Exception:
                end_label = item["date_end"]
            details.append({"tag": "li", "children": [f"🔚  по {end_label}"]})
        p = _fmt_price(item.get("price"), item.get("prepayment"))
        if p != "—":
            details.append({"tag": "li", "children": [f"💰  {p}"]})
        if item.get("confirm_num"):
            details.append({"tag": "li", "children": [f"🔖  {item['confirm_num']}"]})
        if item.get("link"):
            details.append({"tag": "li", "children": [
                {"tag": "a", "attrs": {"href": item["link"]}, "children": ["🔗  Открыть бронирование"]}
            ]})
        if item.get("notes"):
            details.append({"tag": "li", "children": [f"📝  {item['notes']}"]})
        if details:
            result.append({"tag": "ul", "children": details})
        return result

    first = True
    for d in sorted(by_date):
        if not first:
            nodes.append({"tag": "hr"})
        first = False
        try:
            dt = datetime.strptime(d, "%Y-%m-%d")
            day_label = f"📅  {dt.day} {MONTHS_FULL[dt.month]} {dt.year},  {DAYS_FULL[dt.weekday()]}"
        except Exception:
            day_label = f"📅  {d}"
        nodes.append({"tag": "h4", "children": [day_label]})
        for item in by_date[d]:
            nodes.extend(_item_nodes(item))

    if no_date:
        nodes.append({"tag": "hr"})
        nodes.append({"tag": "h4", "children": ["📌  Без даты"]})
        for item in no_date:
            nodes.extend(_item_nodes(item))

    if total_price:
        nodes.append({"tag": "hr"})
        nodes.append({"tag": "blockquote", "children": [
            f"💼 Всего событий: {len(items)}   "
            f"💰 Стоимость: {total_price:,.0f} ₽   "
            f"✅ Оплачено: {total_pre:,.0f} ₽   "
            f"⏳ Осталось: {total_price - total_pre:,.0f} ₽"
        ]})

    return nodes

async def _publish_trip(trip: dict, items: list) -> str:
    token = await _tg_token()
    nodes = _tg_nodes_items(trip["name"], items)
    key_path = f"tg_path_{trip['id']}"
    async with aiosqlite.connect(DB_PATH) as db:
        row = await (await db.execute("SELECT value FROM _meta WHERE key=?", (key_path,))).fetchone()
    page_path = row[0] if row else None
    async with aiohttp.ClientSession() as s:
        ep = f"{TELEGRAPH_API}/{'editPage/' + page_path if page_path else 'createPage'}"
        result = (await (await s.post(ep, json={
            "access_token": token,
            "title": f"Маршрут: {trip['name']}"[:256],
            "content": nodes, "return_content": False
        })).json())["result"]
    if not page_path:
        page_path = result["path"]
        async with aiosqlite.connect(DB_PATH) as db:
            await db.execute("INSERT OR REPLACE INTO _meta VALUES (?,?)", (key_path, page_path))
            await db.commit()
    return f"https://telegra.ph/{page_path}"

@router.message(Command("publish"))
async def cmd_publish(msg: Message):
    if str(msg.from_user.id) not in _load_admins():
        await msg.answer("⛔ Нет доступа"); return
    trip = await _get_active_trip(str(msg.from_user.id))
    if not trip:
        await msg.answer("Нет активного путешествия."); return
    status_msg = await msg.answer("⏳ Публикую...")
    items = await _trip_items(trip["id"])
    if not items:
        await status_msg.edit_text("Данных нет."); return
    url = await _publish_trip(trip, items)
    await status_msg.edit_text(
        f"✅ <b>Опубликовано!</b>\n\n🔗 {url}\n\nОбновить: /publish",
        parse_mode="HTML"
    )

@router.message(Command("weblink"))
async def cmd_weblink(msg: Message):
    trip = await _get_active_trip(str(msg.from_user.id))
    if not trip:
        await msg.answer("Нет активного путешествия."); return
    key_path = f"tg_path_{trip['id']}"
    async with aiosqlite.connect(DB_PATH) as db:
        row = await (await db.execute("SELECT value FROM _meta WHERE key=?", (key_path,))).fetchone()
    if row:
        await msg.answer(
            f"🔗 <b>Онлайн-маршрут:</b>\nhttps://telegra.ph/{row[0]}\n\nОбновить: /publish",
            parse_mode="HTML"
        )
    else:
        await msg.answer("Нажмите /publish для первой публикации.")


# ── ADMIN ─────────────────────────────────────────────────────────────────────

@router.message(Command("addadmin"))
async def cmd_addadmin(msg: Message):
    if str(msg.from_user.id) not in _load_admins(): await msg.answer("⛔ Нет доступа"); return
    parts = msg.text.split()
    if len(parts) < 2 or not parts[1].lstrip("-").isdigit():
        await msg.answer("Использование: /addadmin <id>"); return
    ids = _load_admins(); ids.add(parts[1]); _save_admins(ids)
    await msg.answer(f"✅ <code>{parts[1]}</code> добавлен.", parse_mode="HTML")

@router.message(Command("removeadmin"))
async def cmd_removeadmin(msg: Message):
    if str(msg.from_user.id) not in _load_admins(): await msg.answer("⛔ Нет доступа"); return
    parts = msg.text.split()
    if len(parts) < 2: await msg.answer("Использование: /removeadmin <id>"); return
    ids = _load_admins(); ids.discard(parts[1]); _save_admins(ids)
    await msg.answer(f"✅ <code>{parts[1]}</code> удалён.", parse_mode="HTML")

@router.message(Command("admins"))
async def cmd_admins(msg: Message):
    if str(msg.from_user.id) not in _load_admins(): await msg.answer("⛔ Нет доступа"); return
    ids = _load_admins()
    await msg.answer(
        "👥 " + ("\n".join(f"• <code>{i}</code>" for i in ids) or "Пусто"),
        parse_mode="HTML"
    )


# ── GROUP SUPPORT ─────────────────────────────────────────────────────────────

BOT_DISPLAY_NAME = os.getenv("BOT_DISPLAY_NAME", "").strip()

@router.message(F.chat.type.in_({"group", "supergroup"}), F.text)
async def handle_group_mention(msg: Message):
    if not BOT_DISPLAY_NAME or not msg.text: return
    if msg.from_user and msg.from_user.is_bot: return
    if BOT_DISPLAY_NAME.lower() not in msg.text.lower(): return
    try:
        from anthropic import AsyncAnthropic as _C
        resp = await _C(api_key=os.getenv("ANTHROPIC_API_KEY", "")).messages.create(
            model="claude-haiku-4-5-20251001", max_tokens=300,
            system=f"Ты — {BOT_DISPLAY_NAME}, менеджер маршрутов. Кратко ответь.",
            messages=[{"role": "user", "content": msg.text}]
        )
        await msg.reply(resp.content[0].text)
    except Exception as e:
        logger.error(f"Group mention error: {e}")


# ── MAIN ──────────────────────────────────────────────────────────────────────

async def main():
    bot = Bot(token=os.getenv("BOT_TOKEN"))
    dp = Dispatcher(storage=MemoryStorage())
    dp.include_router(router)
    await bot.set_my_description(BOT_DESCRIPTION)
    await init_db()
    asyncio.create_task(_digest_loop(bot))
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
