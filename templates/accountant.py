# TEMPLATE: accountant
# USE FOR: income/expense tracking, финансы, бухгалтерия, учёт расходов, бюджет
# CUSTOMIZE: sections marked with # CUSTOMIZE
from __future__ import annotations

import asyncio
import json
import logging
import os
import re
from datetime import datetime, date, timedelta
from pathlib import Path

import aiohttp
import aiosqlite
from dotenv import load_dotenv
load_dotenv()
from aiogram import Bot, Dispatcher, F, Router
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import (
    CallbackQuery, FSInputFile, InlineKeyboardButton, InlineKeyboardMarkup,
    KeyboardButton, Message, ReplyKeyboardMarkup,
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ── CUSTOMIZE ────────────────────────────────────────────────────────────────
BOT_DESCRIPTION = "Учёт доходов и расходов по проектам. Категории, баланс, отчёты, Excel-выгрузка."
WELCOME_TEXT = (
    "💼 <b>Финансовый учёт</b>\n\n"
    "Веду все доходы и расходы по проектам:\n"
    "➕ Доходы · ➖ Расходы · 📊 Баланс · 📋 Отчёты\n\n"
    "Выберите действие:"
)
EXPENSE_CATEGORIES = [
    "🏠 Аренда", "🛒 Продукты", "🚗 Транспорт", "💊 Здоровье",
    "🎮 Развлечения", "👔 Одежда", "💡 Коммуналка", "📱 Связь",
    "✈️ Путешествия", "🎓 Обучение", "💰 Другое",
]
INCOME_CATEGORIES = [
    "💵 Зарплата", "💼 Фриланс", "📈 Инвестиции", "🎁 Подарок",
    "🏦 Возврат", "💰 Другой доход",
]
# ── END CUSTOMIZE ─────────────────────────────────────────────────────────────

BOT_NAME = Path(__file__).stem
DATA_DIR = Path(os.getenv("DATA_DIR", "./data"))
DATA_DIR.mkdir(exist_ok=True)
DB_PATH = str(DATA_DIR / f"{BOT_NAME}_data.db")
EXCEL_PATH = str(DATA_DIR / f"{BOT_NAME}_data.xlsx")
HTML_PATH = str(DATA_DIR / f"{BOT_NAME}_report.html")
WELCOME_IMAGE = DATA_DIR / "bot_images" / f"{BOT_NAME}.jpg"
ADMINS_FILE = DATA_DIR / f"admins_{BOT_NAME}.json"
TELEGRAPH_API = "https://api.telegra.ph"

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
router = Router()


# ── admin helpers ─────────────────────────────────────────────────────────────

def _load_admins() -> set:
    try:
        return set(json.loads(ADMINS_FILE.read_text()).get("ids", []))
    except Exception:
        return set()

def _save_admins(ids: set) -> None:
    ADMINS_FILE.write_text(json.dumps({"ids": list(ids)}, ensure_ascii=False))


# ── db ────────────────────────────────────────────────────────────────────────

async def init_db():
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("""
            CREATE TABLE IF NOT EXISTS projects (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                name       TEXT NOT NULL,
                created_at TEXT DEFAULT (datetime('now','localtime'))
            )
        """)
        await db.execute("""
            CREATE TABLE IF NOT EXISTS user_prefs (
                user_id           TEXT PRIMARY KEY,
                active_project_id INTEGER
            )
        """)
        await db.execute("""
            CREATE TABLE IF NOT EXISTS transactions (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL REFERENCES projects(id),
                kind       TEXT NOT NULL CHECK(kind IN ('income','expense')),
                amount     REAL NOT NULL,
                currency   TEXT DEFAULT 'RUB',
                category   TEXT,
                note       TEXT,
                tx_date    TEXT DEFAULT (date('now','localtime')),
                created_at TEXT DEFAULT (datetime('now','localtime'))
            )
        """)
        await db.execute(
            "CREATE TABLE IF NOT EXISTS _meta (key TEXT PRIMARY KEY, value TEXT)"
        )
        # Migration: add project_id to transactions if table was created by an older schema
        try:
            await db.execute("ALTER TABLE transactions ADD COLUMN project_id INTEGER DEFAULT 1")
        except Exception:
            pass  # column already exists
        await db.commit()


# ── project helpers ───────────────────────────────────────────────────────────

async def _all_projects() -> list[dict]:
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        rows = await (await db.execute("SELECT * FROM projects ORDER BY id DESC")).fetchall()
        return [dict(r) for r in rows]

async def _get_active_project(user_id: str) -> dict | None:
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        row = await (await db.execute(
            "SELECT p.* FROM user_prefs u JOIN projects p ON u.active_project_id=p.id WHERE u.user_id=?",
            (user_id,)
        )).fetchone()
        if row:
            return dict(row)
        row = await (await db.execute("SELECT * FROM projects ORDER BY id DESC LIMIT 1")).fetchone()
        return dict(row) if row else None

async def _set_active_project(user_id: str, project_id: int):
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute(
            "INSERT OR REPLACE INTO user_prefs (user_id, active_project_id) VALUES (?,?)",
            (user_id, project_id)
        )
        await db.commit()


MAIN_BUTTONS = frozenset({
    "➕ Доход", "➖ Расход", "💰 Баланс",
    "📋 История", "📊 Отчёт", "📁 Проекты", "📥 Excel", "🌐 HTML",
})

# ── keyboards ─────────────────────────────────────────────────────────────────

def kb_main() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(keyboard=[
        [KeyboardButton(text="➕ Доход"),     KeyboardButton(text="➖ Расход")],
        [KeyboardButton(text="💰 Баланс"),   KeyboardButton(text="📋 История")],
        [KeyboardButton(text="📊 Отчёт"),    KeyboardButton(text="📁 Проекты")],
        [KeyboardButton(text="📥 Excel"),    KeyboardButton(text="🌐 HTML")],
    ], resize_keyboard=True)

def kb_cats(cats: list[str], kind: str) -> InlineKeyboardMarkup:
    rows = [[InlineKeyboardButton(text=c, callback_data=f"cat:{kind}:{c}")] for c in cats]
    rows.append([InlineKeyboardButton(text="❌ Отмена", callback_data="tx_cancel")])
    return InlineKeyboardMarkup(inline_keyboard=rows)

def kb_period() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📅 Сегодня",    callback_data="rperiod:today"),
         InlineKeyboardButton(text="📅 Эта неделя", callback_data="rperiod:week")],
        [InlineKeyboardButton(text="📅 Этот месяц", callback_data="rperiod:month"),
         InlineKeyboardButton(text="📅 Всё время",  callback_data="rperiod:all")],
    ])

def kb_tx_del(tx_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="🗑 Удалить", callback_data=f"tx_del:{tx_id}"),
        InlineKeyboardButton(text="✖ Закрыть", callback_data="tx_close"),
    ]])

def kb_projects(projects: list[dict]) -> InlineKeyboardMarkup:
    rows = [[InlineKeyboardButton(text=f"📁 {p['name']}", callback_data=f"proj_sel:{p['id']}")] for p in projects]
    rows.append([InlineKeyboardButton(text="➕ Новый проект", callback_data="proj_new")])
    rows.append([InlineKeyboardButton(text="❌ Закрыть", callback_data="proj_close")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


# ── FSM ───────────────────────────────────────────────────────────────────────

class AddTx(StatesGroup):
    amount = State(); note = State()

class ProjectCreate(StatesGroup):
    name = State()


# ── /start ────────────────────────────────────────────────────────────────────

@router.message(Command("start"))
async def cmd_start(message: Message):
    admins = _load_admins()
    if not admins:
        _save_admins({str(message.from_user.id)})
    if WELCOME_IMAGE.exists():
        await message.answer_photo(FSInputFile(str(WELCOME_IMAGE)),
                                   caption=WELCOME_TEXT, parse_mode="HTML", reply_markup=kb_main())
    else:
        await message.answer(WELCOME_TEXT, parse_mode="HTML", reply_markup=kb_main())


# ── PROJECTS PANEL ────────────────────────────────────────────────────────────

@router.message(F.text == "📁 Проекты")
async def projects_panel(msg: Message, state: FSMContext):
    await state.clear()
    projects = await _all_projects()
    if not projects:
        await msg.answer(
            "У вас ещё нет проектов.\nСоздайте первый!",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="➕ Новый проект", callback_data="proj_new")],
            ])
        )
        return
    active = await _get_active_project(str(msg.from_user.id))
    header = f"📁 <b>Текущий:</b> {active['name']}\n\n" if active else ""
    await msg.answer(
        header + "Выберите проект или создайте новый:",
        parse_mode="HTML",
        reply_markup=kb_projects(projects)
    )

@router.callback_query(F.data == "proj_new")
async def cb_proj_new(cb: CallbackQuery, state: FSMContext):
    await cb.answer()
    await state.set_state(ProjectCreate.name)
    await cb.message.edit_text(
        "✏️ <b>Введите название проекта</b>\n"
        "Например: Турция 2025, Бизнес, Личные расходы:",
        parse_mode="HTML"
    )

@router.message(ProjectCreate.name, F.text, ~F.text.in_(MAIN_BUTTONS))
async def proj_create_name(msg: Message, state: FSMContext):
    name = msg.text.strip()
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute("INSERT INTO projects (name) VALUES (?)", (name,))
        project_id = cur.lastrowid
        await db.commit()
    await _set_active_project(str(msg.from_user.id), project_id)
    await state.clear()
    await msg.answer(
        f"✅ Создан проект <b>{name}</b>!\n"
        f"Теперь добавляйте доходы и расходы.",
        parse_mode="HTML", reply_markup=kb_main()
    )

@router.callback_query(F.data.startswith("proj_sel:"))
async def cb_proj_sel(cb: CallbackQuery):
    await cb.answer()
    project_id = int(cb.data.split(":")[1])
    async with aiosqlite.connect(DB_PATH) as db:
        row = await (await db.execute("SELECT name FROM projects WHERE id=?", (project_id,))).fetchone()
    if not row:
        await cb.message.edit_text("Проект не найден."); return
    await _set_active_project(str(cb.from_user.id), project_id)
    await cb.message.edit_text(f"✅ Активный проект: <b>{row[0]}</b>", parse_mode="HTML")
    await cb.message.answer("Что сделаем?", reply_markup=kb_main())

@router.callback_query(F.data == "proj_close")
async def cb_proj_close(cb: CallbackQuery):
    await cb.answer()
    await cb.message.delete()

@router.callback_query(F.data == "proj_list")
async def cb_proj_list(cb: CallbackQuery):
    await cb.answer()
    projects = await _all_projects()
    active = await _get_active_project(str(cb.from_user.id))
    header = f"📁 <b>Текущий:</b> {active['name']}\n\n" if active else ""
    await cb.message.edit_text(
        header + "Выберите проект:",
        parse_mode="HTML",
        reply_markup=kb_projects(projects)
    )


# ── ADD INCOME / EXPENSE ──────────────────────────────────────────────────────

@router.message(F.text.in_({"➕ Доход", "➖ Расход"}))
async def add_start(msg: Message, state: FSMContext):
    await state.clear()
    project = await _get_active_project(str(msg.from_user.id))
    if not project:
        await msg.answer("Сначала создайте проект — нажмите 📁 Проекты.", reply_markup=kb_main()); return
    kind = "income" if "Доход" in msg.text else "expense"
    cats = INCOME_CATEGORIES if kind == "income" else EXPENSE_CATEGORIES
    await state.update_data(kind=kind, project_id=project["id"])
    label = "доход" if kind == "income" else "расход"
    await msg.answer(
        f"📁 <b>{project['name']}</b>\n\nВыберите категорию ({label}):",
        parse_mode="HTML",
        reply_markup=kb_cats(cats, kind)
    )

@router.callback_query(F.data.startswith("cat:"))
async def cb_cat(cb: CallbackQuery, state: FSMContext):
    await cb.answer()
    _, kind, cat = cb.data.split(":", 2)
    await state.update_data(category=cat)
    await state.set_state(AddTx.amount)
    label = "дохода" if kind == "income" else "расхода"
    await cb.message.edit_text(f"💰 Введите сумму {label} (цифрами):")

@router.message(AddTx.amount, F.text, ~F.text.in_(MAIN_BUTTONS))
async def add_amount(msg: Message, state: FSMContext):
    cleaned = re.sub(r"[^\d.]", "", msg.text)
    try:
        amount = float(cleaned)
        if amount <= 0: raise ValueError
    except ValueError:
        await msg.answer("Введите положительное число, например: 5000"); return
    await state.update_data(amount=amount)
    await state.set_state(AddTx.note)
    await msg.answer("📝 Комментарий (необязательно):\nОтправьте текст или /skip")

@router.message(AddTx.note, F.text, ~F.text.in_(MAIN_BUTTONS))
async def add_note(msg: Message, state: FSMContext):
    note = None if msg.text.strip() == "/skip" else msg.text.strip()
    await _save_tx(msg, state, note)

@router.message(Command("skip"), AddTx.note)
async def skip_note(msg: Message, state: FSMContext):
    await _save_tx(msg, state, None)

async def _save_tx(msg: Message, state: FSMContext, note):
    data = await state.get_data()
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute(
            "INSERT INTO transactions (project_id, kind, amount, category, note) VALUES (?,?,?,?,?)",
            (data["project_id"], data["kind"], data["amount"], data.get("category"), note)
        )
        await db.commit()
    await state.clear()
    kind_ru = "Доход" if data["kind"] == "income" else "Расход"
    sym = "+" if data["kind"] == "income" else "-"
    await msg.answer(
        f"✅ <b>{kind_ru} записан</b>\n"
        f"💰 {sym}{data['amount']:,.0f} ₽\n"
        f"🏷 {data.get('category', '—')}"
        + (f"\n📝 {note}" if note else ""),
        parse_mode="HTML", reply_markup=kb_main()
    )

@router.callback_query(F.data == "tx_cancel")
async def cb_cancel(cb: CallbackQuery, state: FSMContext):
    await cb.answer(); await state.clear()
    await cb.message.edit_text("Отменено.")
    await cb.message.answer("Что сделаем?", reply_markup=kb_main())


# ── BALANCE ───────────────────────────────────────────────────────────────────

@router.message(F.text == "💰 Баланс")
async def show_balance(msg: Message, state: FSMContext):
    await state.clear()
    project = await _get_active_project(str(msg.from_user.id))
    if not project:
        await msg.answer("Сначала создайте проект — нажмите 📁 Проекты."); return
    pid = project["id"]
    async with aiosqlite.connect(DB_PATH) as db:
        income  = (await (await db.execute(
            "SELECT COALESCE(SUM(amount),0) FROM transactions WHERE project_id=? AND kind='income'", (pid,)
        )).fetchone())[0]
        expense = (await (await db.execute(
            "SELECT COALESCE(SUM(amount),0) FROM transactions WHERE project_id=? AND kind='expense'", (pid,)
        )).fetchone())[0]
        count   = (await (await db.execute(
            "SELECT COUNT(*) FROM transactions WHERE project_id=?", (pid,)
        )).fetchone())[0]
        m_in  = (await (await db.execute(
            "SELECT COALESCE(SUM(amount),0) FROM transactions WHERE project_id=? AND kind='income' "
            "AND strftime('%Y-%m',tx_date)=strftime('%Y-%m','now','localtime')", (pid,)
        )).fetchone())[0]
        m_ex  = (await (await db.execute(
            "SELECT COALESCE(SUM(amount),0) FROM transactions WHERE project_id=? AND kind='expense' "
            "AND strftime('%Y-%m',tx_date)=strftime('%Y-%m','now','localtime')", (pid,)
        )).fetchone())[0]
    balance = income - expense
    sign = "+" if balance >= 0 else ""
    await msg.answer(
        f"💰 <b>Баланс: {project['name']}</b>\n\n"
        f"📈 Доходы (всего): {income:,.0f} ₽\n"
        f"📉 Расходы (всего): {expense:,.0f} ₽\n"
        f"<b>Итого: {sign}{balance:,.0f} ₽</b>\n\n"
        f"<b>Этот месяц:</b>\n"
        f"  📈 +{m_in:,.0f} ₽  📉 -{m_ex:,.0f} ₽\n"
        f"  <b>= {m_in - m_ex:+,.0f} ₽</b>\n\n"
        f"📋 Всего операций: {count}",
        parse_mode="HTML", reply_markup=kb_main()
    )


# ── HISTORY ───────────────────────────────────────────────────────────────────

@router.message(F.text == "📋 История")
async def show_history(msg: Message, state: FSMContext):
    await state.clear()
    project = await _get_active_project(str(msg.from_user.id))
    if not project:
        await msg.answer("Сначала создайте проект — нажмите 📁 Проекты."); return
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        rows = await (await db.execute(
            "SELECT * FROM transactions WHERE project_id=? ORDER BY tx_date DESC, id DESC LIMIT 20",
            (project["id"],)
        )).fetchall()
    if not rows:
        await msg.answer(f"По проекту «{project['name']}» операций нет.", reply_markup=kb_main()); return
    lines = [f"📋 <b>История: {project['name']}</b>\n"]
    for r in [dict(x) for x in rows]:
        sym = "📈" if r["kind"] == "income" else "📉"
        sign = "+" if r["kind"] == "income" else "-"
        lines.append(f"{sym} {r['tx_date']}  <b>{sign}{r['amount']:,.0f} ₽</b>  {r.get('category', '')}")
        if r.get("note"): lines.append(f"   └ {r['note']}")
    btn_rows = [[InlineKeyboardButton(
        text=f"{'📈' if dict(r)['kind'] == 'income' else '📉'} {dict(r)['tx_date']} {dict(r)['amount']:,.0f}₽",
        callback_data=f"tx_view:{dict(r)['id']}"
    )] for r in rows[:10]]
    await msg.answer("\n".join(lines), parse_mode="HTML",
                     reply_markup=InlineKeyboardMarkup(inline_keyboard=btn_rows))

@router.callback_query(F.data.startswith("tx_view:"))
async def cb_tx_view(cb: CallbackQuery):
    await cb.answer()
    tx_id = int(cb.data.split(":")[1])
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        row = await (await db.execute("SELECT * FROM transactions WHERE id=?", (tx_id,))).fetchone()
    if not row:
        await cb.message.answer("Операция не найдена."); return
    row = dict(row)
    sym = "📈 Доход" if row["kind"] == "income" else "📉 Расход"
    sign = "+" if row["kind"] == "income" else "-"
    text = (f"{sym}\n💰 {sign}{row['amount']:,.0f} ₽\n🏷 {row.get('category', '—')}\n"
            f"📅 {row['tx_date']}" + (f"\n📝 {row['note']}" if row.get("note") else ""))
    await cb.message.answer(text, reply_markup=kb_tx_del(tx_id))

@router.callback_query(F.data.startswith("tx_del:"))
async def cb_tx_del(cb: CallbackQuery):
    await cb.answer()
    tx_id = int(cb.data.split(":")[1])
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("DELETE FROM transactions WHERE id=?", (tx_id,)); await db.commit()
    await cb.message.edit_text("🗑 Операция удалена.")

@router.callback_query(F.data == "tx_close")
async def cb_tx_close(cb: CallbackQuery):
    await cb.answer(); await cb.message.delete()


# ── REPORT ────────────────────────────────────────────────────────────────────

@router.message(F.text == "📊 Отчёт")
async def report_start(msg: Message, state: FSMContext):
    await state.clear()
    project = await _get_active_project(str(msg.from_user.id))
    if not project:
        await msg.answer("Сначала создайте проект — нажмите 📁 Проекты."); return
    await msg.answer(
        f"📊 <b>Отчёт: {project['name']}</b>\n\nВыберите период:",
        parse_mode="HTML",
        reply_markup=kb_period()
    )

@router.callback_query(F.data.startswith("rperiod:"))
async def cb_period(cb: CallbackQuery):
    await cb.answer()
    period = cb.data.split(":")[1]
    today = date.today()
    if period == "today":
        d_from = d_to = today.isoformat(); label = "Сегодня"
    elif period == "week":
        d_from = (today - timedelta(days=today.weekday())).isoformat()
        d_to = today.isoformat(); label = "Эта неделя"
    elif period == "month":
        d_from = today.replace(day=1).isoformat(); d_to = today.isoformat(); label = "Этот месяц"
    else:
        d_from = "2000-01-01"; d_to = today.isoformat(); label = "Всё время"

    project = await _get_active_project(str(cb.from_user.id))
    if not project:
        await cb.message.edit_text("Нет активного проекта."); return
    pid = project["id"]

    async with aiosqlite.connect(DB_PATH) as db:
        income = (await (await db.execute(
            "SELECT COALESCE(SUM(amount),0) FROM transactions WHERE project_id=? AND kind='income' AND tx_date BETWEEN ? AND ?",
            (pid, d_from, d_to)
        )).fetchone())[0]
        expense = (await (await db.execute(
            "SELECT COALESCE(SUM(amount),0) FROM transactions WHERE project_id=? AND kind='expense' AND tx_date BETWEEN ? AND ?",
            (pid, d_from, d_to)
        )).fetchone())[0]
        by_cat = await (await db.execute(
            "SELECT category, kind, COALESCE(SUM(amount),0), COUNT(*) FROM transactions "
            "WHERE project_id=? AND tx_date BETWEEN ? AND ? GROUP BY category, kind ORDER BY 3 DESC",
            (pid, d_from, d_to)
        )).fetchall()
        top5_exp = await (await db.execute(
            "SELECT note, amount, tx_date FROM transactions "
            "WHERE project_id=? AND kind='expense' AND tx_date BETWEEN ? AND ? ORDER BY amount DESC LIMIT 5",
            (pid, d_from, d_to)
        )).fetchall()

    lines = [f"📊 <b>Отчёт: {project['name']} — {label}</b>\n",
             f"📈 Доходы: {income:,.0f} ₽",
             f"📉 Расходы: {expense:,.0f} ₽",
             f"<b>Баланс: {income - expense:+,.0f} ₽</b>"]

    if by_cat:
        lines.append("\n<b>По категориям:</b>")
        for cat, kind, amt, cnt in by_cat:
            sym = "📈" if kind == "income" else "📉"
            sign = "+" if kind == "income" else "-"
            lines.append(f"  {sym} {cat or '—'}: {sign}{amt:,.0f} ₽ ({cnt} оп.)")

    if top5_exp:
        lines.append("\n<b>Топ расходов:</b>")
        for note, amt, dt in top5_exp:
            lines.append(f"  • {dt}  {amt:,.0f} ₽" + (f" — {note}" if note else ""))

    cats_with_data = list({row[0] for row in by_cat if row[0]})
    if cats_with_data:
        lines.append("\n<i>Нажмите категорию для детального просмотра:</i>")

    await cb.message.edit_text(
        "\n".join(lines), parse_mode="HTML",
        reply_markup=_kb_cat_filter(cats_with_data, period)
    )

def _kb_cat_filter(cats: list[str], period: str) -> InlineKeyboardMarkup | None:
    if not cats:
        return None
    rows = []
    for i in range(0, len(cats[:8]), 2):
        row = [InlineKeyboardButton(text=cats[i], callback_data=f"cat_filter:{period}:{cats[i]}")]
        if i + 1 < len(cats):
            row.append(InlineKeyboardButton(text=cats[i + 1], callback_data=f"cat_filter:{period}:{cats[i + 1]}"))
        rows.append(row)
    return InlineKeyboardMarkup(inline_keyboard=rows)

@router.callback_query(F.data.startswith("cat_filter:"))
async def cb_cat_filter(cb: CallbackQuery):
    await cb.answer()
    parts = cb.data.split(":", 2)
    period, cat = parts[1], parts[2]
    today = date.today()
    if period == "today":
        d_from = d_to = today.isoformat()
    elif period == "week":
        d_from = (today - timedelta(days=today.weekday())).isoformat(); d_to = today.isoformat()
    elif period == "month":
        d_from = today.replace(day=1).isoformat(); d_to = today.isoformat()
    else:
        d_from = "2000-01-01"; d_to = today.isoformat()

    project = await _get_active_project(str(cb.from_user.id))
    if not project:
        await cb.message.edit_text("Нет активного проекта."); return
    pid = project["id"]

    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        rows = await (await db.execute(
            "SELECT kind, amount, note, tx_date FROM transactions "
            "WHERE project_id=? AND category=? AND tx_date BETWEEN ? AND ? ORDER BY tx_date DESC LIMIT 30",
            (pid, cat, d_from, d_to)
        )).fetchall()
        total = (await (await db.execute(
            "SELECT COALESCE(SUM(amount),0) FROM transactions "
            "WHERE project_id=? AND category=? AND tx_date BETWEEN ? AND ?",
            (pid, cat, d_from, d_to)
        )).fetchone())[0]

    lines = [f"🏷 <b>Категория: {cat}</b>\n📁 {project['name']}\n"]
    for r in [dict(x) for x in rows]:
        sym = "📈" if r["kind"] == "income" else "📉"
        sign = "+" if r["kind"] == "income" else "-"
        line = f"{sym} {r['tx_date']}  <b>{sign}{r['amount']:,.0f} ₽</b>"
        if r.get("note"): line += f"\n   └ {r['note']}"
        lines.append(line)
    lines.append(f"\n<b>Итого: {total:,.0f} ₽</b>")

    await cb.message.edit_text(
        "\n".join(lines), parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="◀️ Назад к отчёту", callback_data=f"rperiod:{period}")
        ]])
    )


# ── EXCEL ─────────────────────────────────────────────────────────────────────

@router.message(F.text == "📥 Excel")
@router.message(Command("excel"))
async def cmd_excel(msg: Message, state: FSMContext):
    await state.clear()
    if str(msg.from_user.id) not in _load_admins():
        await msg.answer("⛔ Нет доступа"); return
    projects = await _all_projects()
    if not projects:
        await msg.answer("Данных нет."); return
    wb = Workbook()
    wb.remove(wb.active)
    hdrs = ["Тип", "Сумма", "Категория", "Комментарий", "Дата", "Создано"]
    hfill = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(bold=True, color="FFFFFF")
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"),  bottom=Side(style="thin"))
    fills = [PatternFill("solid", fgColor="FFFFFF"), PatternFill("solid", fgColor="DCE6F1")]
    green = Font(color="155724"); red = Font(color="721C24")
    total_fill = PatternFill("solid", fgColor="FFF2CC")
    total_font = Font(bold=True)

    for project in projects:
        async with aiosqlite.connect(DB_PATH) as db:
            db.row_factory = aiosqlite.Row
            rows = [dict(r) for r in await (await db.execute(
                "SELECT * FROM transactions WHERE project_id=? ORDER BY tx_date, id",
                (project["id"],)
            )).fetchall()]
        ws = wb.create_sheet(title=project["name"][:31])
        for c, h in enumerate(hdrs, 1):
            cell = ws.cell(1, c, h); cell.fill = hfill; cell.font = hfont; cell.border = border
            cell.alignment = Alignment(horizontal="center")
        total_income = 0; total_expense = 0
        for ri, r in enumerate(rows, 2):
            kind_ru = "Доход" if r["kind"] == "income" else "Расход"
            vals = [kind_ru, r["amount"], r.get("category", ""), r.get("note", ""),
                    r["tx_date"], r["created_at"]]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(ri, c, v); cell.fill = fills[ri % 2]; cell.border = border
                if c == 2:
                    cell.font = green if r["kind"] == "income" else red
            if r["kind"] == "income":
                total_income += r["amount"]
            else:
                total_expense += r["amount"]
        tr = len(rows) + 2
        ws.cell(tr, 1, "ИТОГО").font = total_font; ws.cell(tr, 1).fill = total_fill
        ws.cell(tr, 2, f"+{total_income:,.0f} / -{total_expense:,.0f}").font = total_font
        ws.cell(tr, 2).fill = total_fill
        ws.cell(tr, 3, f"Баланс: {total_income - total_expense:+,.0f} ₽").font = total_font
        ws.cell(tr, 3).fill = total_fill
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(
                max(max(len(str(c.value or "")) for c in col) + 2, 10), 40
            )
        ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:F{len(rows) + 1}"

    wb.save(EXCEL_PATH)
    await msg.answer_document(FSInputFile(EXCEL_PATH),
                              caption=f"📊 Финансовая выгрузка ({len(projects)} проектов)")


# ── HTML REPORT ───────────────────────────────────────────────────────────────

_HTML_REPORT = r"""<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Финансовый отчёт</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
:root{--bg:#0d1117;--card:#161b22;--border:#30363d;--text:#e6edf3;--muted:#8b949e;--green:#3fb950;--red:#f85149;--blue:#58a6ff}
*{margin:0;padding:0;box-sizing:border-box}
body{background:var(--bg);color:var(--text);font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;padding:20px;min-height:100vh}
h1{font-size:1.5rem;margin-bottom:4px}
.sub{color:var(--muted);font-size:.85rem;margin-bottom:24px}
.stats{display:grid;grid-template-columns:repeat(3,1fr);gap:14px;margin-bottom:24px}
.stat-card{background:var(--card);border:1px solid var(--border);border-radius:14px;padding:18px}
.stat-card .lbl{color:var(--muted);font-size:.75rem;text-transform:uppercase;letter-spacing:.06em;margin-bottom:8px}
.stat-card .val{font-size:1.6rem;font-weight:700}
.c-green{color:var(--green)}.c-red{color:var(--red)}.c-blue{color:var(--blue)}
.filters{background:var(--card);border:1px solid var(--border);border-radius:14px;padding:16px;margin-bottom:24px;display:flex;flex-wrap:wrap;gap:14px;align-items:flex-end}
.fg{display:flex;flex-direction:column;gap:6px}
.fg label{color:var(--muted);font-size:.75rem;text-transform:uppercase;letter-spacing:.06em}
select,input[type=date]{background:#21262d;color:var(--text);border:1px solid var(--border);border-radius:8px;padding:7px 10px;font-size:.88rem;outline:none}
.tbtn-row,.pbtn-row{display:flex;gap:5px;flex-wrap:wrap}
.tbtn,.pbtn{background:#21262d;color:var(--muted);border:1px solid var(--border);border-radius:8px;padding:7px 14px;font-size:.82rem;cursor:pointer;transition:all .15s}
.tbtn.active,.pbtn.active{background:var(--blue);color:#fff;border-color:var(--blue)}
.chart-wrap{background:var(--card);border:1px solid var(--border);border-radius:14px;padding:20px;margin-bottom:24px;display:flex;align-items:center;gap:28px;flex-wrap:wrap}
.chart-wrap canvas{max-width:190px;max-height:190px}
.legend{display:flex;flex-direction:column;gap:10px}
.leg-item{display:flex;align-items:center;gap:9px;font-size:.88rem}
.leg-dot{width:11px;height:11px;border-radius:50%;flex-shrink:0}
.table-wrap{background:var(--card);border:1px solid var(--border);border-radius:14px;overflow:hidden}
table{width:100%;border-collapse:collapse}
th{background:#1c2128;color:var(--muted);font-size:.75rem;text-transform:uppercase;letter-spacing:.06em;padding:11px 14px;text-align:left;cursor:pointer;user-select:none;white-space:nowrap}
th:hover{color:var(--text)}
th.srt .arr{opacity:1!important}
.arr{margin-left:4px;opacity:.3}
td{padding:11px 14px;border-top:1px solid var(--border);font-size:.88rem;vertical-align:middle}
tr:hover td{background:#1c2128}
.badge{display:inline-block;padding:2px 9px;border-radius:12px;font-size:.75rem;font-weight:600}
.bi{background:#1a3a1a;color:var(--green)}.be{background:#3a1a1a;color:var(--red)}
.ai{color:var(--green);font-weight:600}.ae{color:var(--red);font-weight:600}
.chip{background:#21262d;border:1px solid var(--border);border-radius:6px;padding:2px 8px;font-size:.75rem;color:var(--muted)}
.empty{text-align:center;padding:40px;color:var(--muted)}
.foot{display:flex;justify-content:space-between;align-items:center;padding:12px 0;color:var(--muted);font-size:.82rem;margin-top:8px}
.csv-btn{background:#21262d;color:var(--text);border:1px solid var(--border);border-radius:8px;padding:6px 14px;font-size:.82rem;cursor:pointer}
.csv-btn:hover{background:#30363d}
#custom-rng{display:none}
@media(max-width:580px){.stats{grid-template-columns:1fr 1fr}.stat-card:last-child{grid-column:span 2}}
</style>
</head>
<body>
<h1>📊 Финансовый отчёт</h1>
<div class="sub" id="sub"></div>
<div class="stats">
  <div class="stat-card"><div class="lbl">Баланс</div><div class="val c-blue" id="s-bal">—</div></div>
  <div class="stat-card"><div class="lbl">Доходы</div><div class="val c-green" id="s-inc">—</div></div>
  <div class="stat-card"><div class="lbl">Расходы</div><div class="val c-red" id="s-exp">—</div></div>
</div>
<div class="filters">
  <div class="fg"><label>Проект</label><select id="fp"><option value="">Все проекты</option></select></div>
  <div class="fg"><label>Тип</label>
    <div class="tbtn-row">
      <button class="tbtn active" data-t="">Все</button>
      <button class="tbtn" data-t="income">Доход</button>
      <button class="tbtn" data-t="expense">Расход</button>
    </div>
  </div>
  <div class="fg"><label>Категория</label><select id="fc"><option value="">Все категории</option></select></div>
  <div class="fg"><label>Период</label>
    <div class="pbtn-row">
      <button class="pbtn" data-p="today">Сегодня</button>
      <button class="pbtn" data-p="week">Неделя</button>
      <button class="pbtn" data-p="month">Месяц</button>
      <button class="pbtn" data-p="year">Год</button>
      <button class="pbtn active" data-p="all">Всё</button>
      <button class="pbtn" data-p="custom">Диапазон</button>
    </div>
  </div>
  <div class="fg" id="custom-rng"><label>Диапазон</label>
    <div style="display:flex;gap:6px;align-items:center">
      <input type="date" id="fd1"><span style="color:var(--muted)">—</span><input type="date" id="fd2">
    </div>
  </div>
</div>
<div class="chart-wrap"><canvas id="chart"></canvas><div class="legend" id="legend"></div></div>
<div class="table-wrap">
<table>
  <thead><tr>
    <th data-c="tx_date">Дата<span class="arr">↕</span></th>
    <th data-c="kind">Тип<span class="arr">↕</span></th>
    <th data-c="amount">Сумма<span class="arr">↕</span></th>
    <th data-c="category">Категория<span class="arr">↕</span></th>
    <th data-c="note">Комментарий<span class="arr">↕</span></th>
    <th data-c="_proj">Проект<span class="arr">↕</span></th>
  </tr></thead>
  <tbody id="tb"></tbody>
</table>
</div>
<div class="foot"><span id="cnt"></span><button class="csv-btn" onclick="dl()">⬇ CSV</button></div>
<script>
const TXS=__DATA_JSON__;
const PROJS=__PROJECTS_JSON__;
document.getElementById('sub').textContent='Сформирован: __GENERATED_AT__';
const pm={};PROJS.forEach(p=>pm[p.id]=p.name);
TXS.forEach(t=>t._proj=pm[t.project_id]||'—');
const sel=document.getElementById('fp');
PROJS.forEach(p=>{const o=document.createElement('option');o.value=p.id;o.textContent=p.name;sel.appendChild(o);});
const sc=document.getElementById('fc');
[...new Set(TXS.map(t=>t.category).filter(Boolean))].sort().forEach(c=>{const o=document.createElement('option');o.value=c;o.textContent=c;sc.appendChild(o);});
let fP='',fT='',fC='',fPer='all',fd1='',fd2='',sCol='tx_date',sDir=-1;
function filtered(){
  const now=new Date();
  return TXS.filter(t=>{
    if(fP&&String(t.project_id)!==String(fP))return false;
    if(fT&&t.kind!==fT)return false;
    if(fC&&t.category!==fC)return false;
    if(fPer!=='all'&&fPer!=='custom'){
      const d=new Date(t.tx_date);
      if(fPer==='today'&&d.toDateString()!==now.toDateString())return false;
      if(fPer==='week'){const w=new Date(now);w.setDate(now.getDate()-7);if(d<w)return false;}
      if(fPer==='month'){const m=new Date(now);m.setMonth(now.getMonth()-1);if(d<m)return false;}
      if(fPer==='year'){const y=new Date(now);y.setFullYear(now.getFullYear()-1);if(d<y)return false;}
    }
    if(fPer==='custom'){if(fd1&&t.tx_date<fd1)return false;if(fd2&&t.tx_date>fd2)return false;}
    return true;
  });
}
const fmt=n=>new Intl.NumberFormat('ru-RU').format(Math.abs(n))+' ₽';
const pal=['#58a6ff','#3fb950','#d29922','#f85149','#bc8cff','#ffa657','#39d353','#ff7b72'];
let ch=null;
function render(){
  const data=filtered().slice().sort((a,b)=>{
    let av=a[sCol]??'',bv=b[sCol]??'';
    return(typeof av==='number'?(av-bv):String(av).localeCompare(String(bv)))*sDir;
  });
  let inc=0,exp=0;data.forEach(t=>t.kind==='income'?inc+=t.amount:exp+=t.amount);
  const bal=inc-exp;
  document.getElementById('s-bal').textContent=(bal>=0?'+':'')+fmt(bal);
  document.getElementById('s-inc').textContent='+'+fmt(inc);
  document.getElementById('s-exp').textContent='-'+fmt(exp);
  const cd={};data.forEach(t=>{const k=t.category||'Без категории';cd[k]=(cd[k]||0)+t.amount;});
  const cl=Object.keys(cd),cv=Object.values(cd),cc=cl.map((_,i)=>pal[i%pal.length]);
  if(ch)ch.destroy();
  ch=new Chart(document.getElementById('chart'),{
    type:'doughnut',data:{labels:cl,datasets:[{data:cv,backgroundColor:cc,borderWidth:0,hoverOffset:4}]},
    options:{cutout:'68%',plugins:{legend:{display:false},tooltip:{callbacks:{label:c=>` ${c.label}: ${fmt(c.raw)}`}}},animation:{duration:250}}
  });
  document.getElementById('legend').innerHTML=cl.map((l,i)=>`<div class="leg-item"><div class="leg-dot" style="background:${cc[i]}"></div><span>${l} — ${fmt(cv[i])}</span></div>`).join('');
  const esc=s=>String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;');
  document.getElementById('tb').innerHTML=data.length?data.map(t=>`<tr>
    <td>${esc(t.tx_date)}</td>
    <td><span class="badge ${t.kind==='income'?'bi':'be'}">${t.kind==='income'?'Доход':'Расход'}</span></td>
    <td class="${t.kind==='income'?'ai':'ae'}">${t.kind==='income'?'+':'-'}${fmt(t.amount)}</td>
    <td>${esc(t.category)}</td><td>${esc(t.note)}</td>
    <td><span class="chip">${esc(t._proj)}</span></td></tr>`).join('')
    :'<tr><td colspan="6" class="empty">Нет данных</td></tr>';
  document.getElementById('cnt').textContent=`Записей: ${data.length}`;
  document.querySelectorAll('th[data-c]').forEach(th=>{
    th.classList.toggle('srt',th.dataset.c===sCol);
    th.querySelector('.arr').textContent=th.dataset.c===sCol?(sDir===1?'↑':'↓'):'↕';
  });
}
document.getElementById('fp').addEventListener('change',e=>{fP=e.target.value;render();});
document.getElementById('fc').addEventListener('change',e=>{fC=e.target.value;render();});
document.getElementById('fd1').addEventListener('change',e=>{fd1=e.target.value;render();});
document.getElementById('fd2').addEventListener('change',e=>{fd2=e.target.value;render();});
document.querySelectorAll('.tbtn').forEach(b=>b.addEventListener('click',()=>{
  document.querySelectorAll('.tbtn').forEach(x=>x.classList.remove('active'));
  b.classList.add('active');fT=b.dataset.t;render();
}));
document.querySelectorAll('.pbtn').forEach(b=>b.addEventListener('click',()=>{
  document.querySelectorAll('.pbtn').forEach(x=>x.classList.remove('active'));
  b.classList.add('active');fPer=b.dataset.p;
  document.getElementById('custom-rng').style.display=fPer==='custom'?'':'none';
  render();
}));
document.querySelectorAll('th[data-c]').forEach(th=>th.addEventListener('click',()=>{
  sCol===th.dataset.c?sDir*=-1:(sCol=th.dataset.c,sDir=-1);render();
}));
function dl(){
  const rows=[['Дата','Тип','Сумма','Категория','Комментарий','Проект'],
    ...filtered().map(t=>[t.tx_date,t.kind==='income'?'Доход':'Расход',t.amount,t.category||'',t.note||'',t._proj])];
  const csv=rows.map(r=>r.map(v=>`"${String(v).replace(/"/g,'""')}"`).join(',')).join('\n');
  const a=document.createElement('a');a.href='data:text/csv;charset=utf-8,﻿'+encodeURIComponent(csv);
  a.download='report.csv';a.click();
}
render();
</script>
</body>
</html>"""


def _build_html(transactions: list, projects: list) -> str:
    from datetime import datetime as _dt
    return (
        _HTML_REPORT
        .replace("__DATA_JSON__", json.dumps(transactions, ensure_ascii=False))
        .replace("__PROJECTS_JSON__", json.dumps(projects, ensure_ascii=False))
        .replace("__GENERATED_AT__", _dt.now().strftime("%d.%m.%Y %H:%M"))
    )


@router.message(F.text == "🌐 HTML")
@router.message(Command("html"))
async def cmd_html_report(msg: Message, state: FSMContext):
    await state.clear()
    if str(msg.from_user.id) not in _load_admins():
        await msg.answer("⛔ Нет доступа"); return
    projects = await _all_projects()
    if not projects:
        await msg.answer("Данных нет."); return
    all_txs: list = []
    for project in projects:
        async with aiosqlite.connect(DB_PATH) as db:
            db.row_factory = aiosqlite.Row
            rows = await (await db.execute(
                "SELECT * FROM transactions WHERE project_id=? ORDER BY tx_date, id",
                (project["id"],)
            )).fetchall()
            all_txs.extend(dict(r) for r in rows)
    html = _build_html(all_txs, projects)
    Path(HTML_PATH).write_text(html, encoding="utf-8")
    await msg.answer_document(
        FSInputFile(HTML_PATH),
        caption="🌐 Открой файл в браузере — фильтры, диаграмма и сортировка внутри"
    )


# ── TELEGRAPH ─────────────────────────────────────────────────────────────────

async def _tg_token() -> str:
    async with aiosqlite.connect(DB_PATH) as db:
        row = await (await db.execute("SELECT value FROM _meta WHERE key='tg_token'")).fetchone()
    if row: return row[0]
    async with aiohttp.ClientSession() as s:
        async with s.post(f"{TELEGRAPH_API}/createAccount",
                          data={"short_name": BOT_NAME[:31], "author_name": "AccountantBot"}) as r:
            token = (await r.json())["result"]["access_token"]
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("INSERT OR REPLACE INTO _meta VALUES ('tg_token',?)", (token,)); await db.commit()
    return token

async def _publish_project(project: dict, rows: list[dict]) -> str:
    token = await _tg_token()
    nodes = [{"tag": "h3", "children": [f"Финансы: {project['name']}"]}]
    for r in rows[:100]:
        sym = "+" if r["kind"] == "income" else "-"
        parts = [f"{sym}{r['amount']:,.0f} ₽  {r.get('category', '')}  {r['tx_date']}"]
        if r.get("note"): parts.append(f"— {r['note']}")
        nodes.append({"tag": "p", "children": [" ".join(parts)]})
    total_in  = sum(r["amount"] for r in rows if r["kind"] == "income")
    total_exp = sum(r["amount"] for r in rows if r["kind"] == "expense")
    nodes.append({"tag": "p", "children": [
        f"Доходы: +{total_in:,.0f} ₽  |  Расходы: -{total_exp:,.0f} ₽  |  Баланс: {total_in - total_exp:+,.0f} ₽"
    ]})
    key_path = f"tg_path_{project['id']}"
    async with aiosqlite.connect(DB_PATH) as db:
        row = await (await db.execute("SELECT value FROM _meta WHERE key=?", (key_path,))).fetchone()
    page_path = row[0] if row else None
    async with aiohttp.ClientSession() as s:
        ep = f"{TELEGRAPH_API}/{'editPage/' + page_path if page_path else 'createPage'}"
        result = (await (await s.post(ep, json={
            "access_token": token,
            "title": f"Финансы: {project['name']}"[:256],
            "content": nodes,
            "return_content": False
        })).json())["result"]
    if not page_path:
        page_path = result["path"]
        async with aiosqlite.connect(DB_PATH) as db:
            await db.execute("INSERT OR REPLACE INTO _meta VALUES (?,?)", (key_path, page_path))
            await db.commit()
    return f"https://telegra.ph/{page_path}"

@router.message(Command("publish"))
async def cmd_publish(msg: Message):
    if str(msg.from_user.id) not in _load_admins():
        await msg.answer("⛔ Нет доступа"); return
    project = await _get_active_project(str(msg.from_user.id))
    if not project:
        await msg.answer("Нет активного проекта."); return
    status_msg = await msg.answer("⏳ Публикую...")
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        rows = [dict(r) for r in await (await db.execute(
            "SELECT * FROM transactions WHERE project_id=? ORDER BY tx_date DESC LIMIT 200",
            (project["id"],)
        )).fetchall()]
    if not rows:
        await status_msg.edit_text("Данных нет."); return
    url = await _publish_project(project, rows)
    await status_msg.edit_text(
        f"✅ <b>Опубликовано!</b>\n\n🔗 {url}\n\nОбновить: /publish", parse_mode="HTML"
    )

@router.message(Command("weblink"))
async def cmd_weblink(msg: Message):
    project = await _get_active_project(str(msg.from_user.id))
    if not project:
        await msg.answer("Нет активного проекта."); return
    key_path = f"tg_path_{project['id']}"
    async with aiosqlite.connect(DB_PATH) as db:
        row = await (await db.execute("SELECT value FROM _meta WHERE key=?", (key_path,))).fetchone()
    if row:
        url = f"https://telegra.ph/{row[0]}"
        await msg.answer(f"🔗 <b>Финансовый отчёт онлайн:</b>\n{url}\n\nОбновить: /publish", parse_mode="HTML")
    else:
        await msg.answer("Нажмите /publish для первой публикации.")


# ── ADMIN ─────────────────────────────────────────────────────────────────────

@router.message(Command("addadmin"))
async def cmd_addadmin(msg: Message):
    if str(msg.from_user.id) not in _load_admins(): await msg.answer("⛔ Нет доступа"); return
    parts = msg.text.split()
    if len(parts) < 2 or not parts[1].lstrip("-").isdigit(): await msg.answer("Использование: /addadmin <id>"); return
    ids = _load_admins(); ids.add(parts[1]); _save_admins(ids)
    await msg.answer(f"✅ <code>{parts[1]}</code> добавлен.", parse_mode="HTML")

@router.message(Command("removeadmin"))
async def cmd_removeadmin(msg: Message):
    if str(msg.from_user.id) not in _load_admins(): await msg.answer("⛔ Нет доступа"); return
    parts = msg.text.split()
    if len(parts) < 2: await msg.answer("Использование: /removeadmin <id>"); return
    ids = _load_admins(); ids.discard(parts[1]); _save_admins(ids)
    await msg.answer(f"✅ <code>{parts[1]}</code> удалён.", parse_mode="HTML")

@router.message(Command("admins"))
async def cmd_admins(msg: Message):
    if str(msg.from_user.id) not in _load_admins(): await msg.answer("⛔ Нет доступа"); return
    ids = _load_admins()
    await msg.answer("👥 " + ("\n".join(f"• <code>{i}</code>" for i in ids) or "Пусто"), parse_mode="HTML")


# ── MAIN ──────────────────────────────────────────────────────────────────────

async def main():
    bot = Bot(token=os.getenv("BOT_TOKEN"))
    dp = Dispatcher(storage=MemoryStorage())
    dp.include_router(router)
    await bot.set_my_description(BOT_DESCRIPTION)
    await init_db()
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
