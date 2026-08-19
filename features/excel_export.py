# FEATURE: excel_export
# COMPATIBLE_WITH: accountant, booking_beauty, booking_fitness, booking_medical, booking_restaurant, campaign_tracker, car_rental, coworking_space, debtors, delivery_tracker, event_manager, event_rsvp, expense_tracker, feedback_survey, habit_tracker, inventory, loyalty_program, manager_secretary, moderator, orders_tracker, referral_program, rental_equipment, repair_tracker, shop_catalog, staff_scheduler, support_tickets, tour_operator, tourist_documents, trip_manager, vehicle_service
"""Generic "export my current data to Excel" feature — a formatted .xlsx
snapshot of one bot's own records, built with openpyxl.

Two halves, deliberately kept separate:

  1. `fetch_export_rows()` — data ACCESS. Modeled directly on
     features/sales_analytics.py's compute_metrics()/_resolve_config(): reads
     office_hook_config (db/database.py's bot_office_hook_config, produced by
     services/claude_service.py's _generate_office_hook_config) to find which
     table is this bot's "own business data", re-validates that table/its
     columns against the bot's ACTUAL sqlite schema (PRAGMA table_info) at
     call time — never trusting office_hook_config blindly, since it can be
     stale after a /recreate that changed the schema — and only ever
     interpolates identifiers that passed the same strict
     `^[A-Za-z_][A-Za-z0-9_]*$` allowlist sales_analytics.py uses, always
     double-quoted. Same "degrade to unavailable, never partially wrong"
     contract: returns None rather than raising when there's nothing safe to
     export, exactly like compute_metrics() returning None.

  2. `build_workbook()` — pure formatting. Modeled on features/word_export.py's
     build_document(): a plain sync function, no I/O, no database, taking
     already-fetched rows and returning raw .xlsx bytes via an in-memory
     BytesIO. openpyxl's cell/style writes for a bounded (MAX_ROWS-capped)
     sheet are well under a millisecond for realistic sizes, so — same
     reasoning word_export.py gives for python-docx — this is NOT wrapped in
     asyncio.to_thread.

Delivery — the one place this feature genuinely differs from both of its
models — needed a decision sales_analytics.py never had to make (it only ever
returns JSON, nothing downloadable) and word_export.py sidesteps (it's called
programmatically from a host template's own chat-command handler, which
already has a Bot instance and chat_id to hand build_document()'s bytes
straight to answer_document()).

Chose mini-app delivery, NOT a chat command, and NOT a router here:
  - This reads office_hook_config the exact same way sales_analytics.py does,
    which is a mini-app-only feature by explicit product-owner decision (see
    that file's docstring) — "export MY current data" is naturally the same
    surface as "view MY current data" (the analytics screen), not a new
    Telegram command surface. A bot owner already on the analytics screen in
    the SPA tapping "Export to Excel" is the natural flow; a brand new /export
    chat command would duplicate that entry point for no benefit.
  - Unlike sales_analytics.py's JSON-only analytics_handler, THIS payload is
    a binary file, so it needs its own aiohttp route returning
    application/vnd.openxmlformats-officedocument.spreadsheetml.sheet bytes
    with a Content-Disposition header — see export_handler() in
    runtime/miniapp_api.py (wired in register_routes() the same way
    analytics_handler is, reserved ahead of the generic {resource} route).
  - No `router` attribute here (same as sales_analytics.py) — the
    runtime/registry.py loader tolerates that via getattr(module, "router",
    None) exactly as documented in that file's own docstring; no registry.py
    edit was needed to add this feature. No `init_db` either — this module
    owns no tables of its own, only reads what office_hook_config points at.

  A future host template's chat command CAN still call build_workbook()
  directly and send the bytes via BufferedInputFile (same shape as
  word_export.py's callers) if a product decision later wants that surface
  too — nothing here precludes it, the two halves are independent by design.
"""
from __future__ import annotations

import io
import logging
import re
from datetime import datetime, date
from typing import Any

import aiosqlite
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

_IDENTIFIER_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")

# ── module-level CUSTOMIZE constants (this codebase's convention for
# per-feature tunables — see e.g. features/sellable_items.py's
# NAME_MAX_LEN/PRICE_MAX/FLOW_TIMEOUT_SECONDS block — not env vars) ─────────

# Hard cap on rows written per sheet. Protects both the export itself (an
# unbounded SELECT * against a table that has grown large over a bot's
# lifetime) and the eventual Telegram/HTTP transfer of the resulting file.
# 10k rows of a typical ~8-column business record table stays comfortably
# under MAX_FILE_SIZE below even with styling applied to every cell.
MAX_ROWS = 10_000

# Guard against ever trying to build/return a workbook that's absurdly large
# (e.g. very wide rows with long text columns even under MAX_ROWS). Checked
# AFTER the workbook is built (openpyxl has no cheap way to predict final
# .xlsx size before saving) — build_workbook() logs and still returns the
# bytes; it is the CALLER's job to decide whether to actually deliver a file
# that exceeds this, since what's "too large to send" depends on the delivery
# channel (Telegram's own bot API cap is 50MB for sendDocument, matching this
# default).
MAX_FILE_SIZE = 50 * 1024 * 1024

HEADER_COLOR = "4472C4"  # Excel's own default "blue, accent 1" — familiar to anyone who's used a real spreadsheet
HEADER_FONT_COLOR = "FFFFFF"
BORDER_STYLE = "thin"

# openpyxl column widths are in "character units" of the default font, not
# pixels — this multiplier converts a computed max string length into a width
# that comfortably fits it (a bare 1:1 ratio visually clips wide fonts /
# unicode text, e.g. Cyrillic labels used throughout this codebase's own
# templates).
AUTO_WIDTH_MULTIPLIER = 1.2
_MIN_COLUMN_WIDTH = 8
_MAX_COLUMN_WIDTH = 60

# Status-like column names get conditional fill colors keyed by value (case-
# insensitive substring match) — a light, best-effort touch, not a general
# rules engine. Columns are matched by NAME (any column whose name contains
# one of these tokens), values by their own lowercased text.
_STATUS_COLUMN_HINTS = ("status", "статус", "state", "active")
_STATUS_COLORS: dict[str, str] = {
    "done": "C6EFCE", "completed": "C6EFCE", "paid": "C6EFCE", "active": "C6EFCE",
    "success": "C6EFCE", "оплачен": "C6EFCE", "завершен": "C6EFCE", "активен": "C6EFCE",
    "pending": "FFEB9C", "waiting": "FFEB9C", "в процессе": "FFEB9C", "ожидание": "FFEB9C",
    "cancelled": "FFC7CE", "canceled": "FFC7CE", "failed": "FFC7CE", "inactive": "FFC7CE",
    "отменен": "FFC7CE", "отменён": "FFC7CE", "неактивен": "FFC7CE",
}


# ── data access — same posture as features/sales_analytics.py ──────────────

async def _table_columns(db: aiosqlite.Connection, table: str) -> list[str] | None:
    """Same helper/rationale as sales_analytics.py's own _table_columns — kept
    as a separate copy rather than a shared import (no shared features/_common
    module exists in this project; every feature independently defines these
    tiny helpers, same convention noted in sellable_items.py). Returns column
    names in schema order (unlike sales_analytics.py's set, order matters here
    since it becomes the sheet's header row order) or None if the table
    doesn't exist."""
    async with db.execute(f'PRAGMA table_info("{table}")') as cursor:
        rows = await cursor.fetchall()
    if not rows:
        return None
    return [row[1] for row in rows]


def _resolve_table(hook_config: dict[str, Any] | None) -> str | None:
    """Extracts and identifier-validates just the table name from a raw
    office_hook_config dict — unlike sales_analytics.py's _resolve_config,
    match_field/created_at_field aren't needed here: an export wants EVERY
    column of the record, not a single grouping/time field."""
    if not hook_config:
        return None
    table = hook_config.get("table")
    if not isinstance(table, str) or not _IDENTIFIER_RE.match(table):
        return None
    return table


async def fetch_export_rows(
    db_path: str, hook_config: dict[str, Any] | None, *, max_rows: int = MAX_ROWS
) -> tuple[list[str], list[list[Any]]] | None:
    """Returns (column_names, rows) for this bot's own business-data table, as
    named by office_hook_config — or None if there's nothing safe to export
    (no usable office_hook_config, or its table no longer exists in this
    bot's ACTUAL schema right now — same stale-config defense as
    sales_analytics.py's compute_metrics()).

    Never raises — any failure degrades to None, same contract as
    compute_metrics(). Rows are capped at `max_rows` (MAX_ROWS by default),
    newest-first when the table has an `id` column (the common autoincrement
    primary key across this codebase's templates) so a truncated export still
    shows the most RECENT records rather than silently dropping them in favor
    of the oldest — falls back to insertion order (no ORDER BY) when there's
    no `id` column to sort by."""
    table = _resolve_table(hook_config)
    if table is None:
        return None

    try:
        async with aiosqlite.connect(db_path) as db:
            columns = await _table_columns(db, table)
            if columns is None:
                logger.warning(f"fetch_export_rows: table {table!r} from office_hook_config no longer exists — stale config")
                return None

            order_clause = ' ORDER BY "id" DESC' if "id" in columns else ""
            column_list = ", ".join(f'"{c}"' for c in columns)
            async with db.execute(
                f'SELECT {column_list} FROM "{table}"{order_clause} LIMIT ?', (max_rows,)
            ) as cursor:
                rows = await cursor.fetchall()
    except Exception:
        logger.warning(f"fetch_export_rows: failed for table={table!r}", exc_info=True)
        return None

    return columns, [list(row) for row in rows]


# ── formatting — same posture as features/word_export.py ───────────────────

def _cell_display_value(value: Any) -> Any:
    """openpyxl accepts str/int/float/bool/datetime/None directly, but sqlite
    can also hand back bytes (a BLOB column) or other exotic types no
    template in this codebase currently stores in an office_hook_config
    table, but which a stale/hand-edited config could still point at —
    stringified defensively rather than letting Workbook.save() raise."""
    if value is None or isinstance(value, (str, int, float, bool, datetime, date)):
        return value
    if isinstance(value, bytes):
        return f"<binary: {len(value)} bytes>"
    return str(value)


def _status_fill_for(column_name: str, value: Any) -> PatternFill | None:
    lowered_column = column_name.lower()
    if not any(hint in lowered_column for hint in _STATUS_COLUMN_HINTS):
        return None
    if value is None:
        return None
    lowered_value = str(value).strip().lower()
    for token, color in _STATUS_COLORS.items():
        if token in lowered_value:
            return PatternFill(start_color=color, end_color=color, fill_type="solid")
    return None


def _autosize_columns(sheet: Worksheet, columns: list[str], rows: list[list[Any]]) -> None:
    for col_idx, column_name in enumerate(columns, start=1):
        longest = len(str(column_name))
        for row in rows:
            value = row[col_idx - 1] if col_idx - 1 < len(row) else None
            longest = max(longest, len(str(value)) if value is not None else 0)
        width = max(_MIN_COLUMN_WIDTH, min(_MAX_COLUMN_WIDTH, int(longest * AUTO_WIDTH_MULTIPLIER) + 2))
        sheet.column_dimensions[get_column_letter(col_idx)].width = width


def build_workbook(
    columns: list[str],
    rows: list[list[Any]],
    *,
    sheet_title: str = "Export",
    max_rows: int = MAX_ROWS,
) -> bytes:
    """Builds a styled .xlsx workbook from already-fetched (columns, rows) —
    the same "pure function over plain data, no I/O" shape as
    word_export.py's build_document(). Truncates to `max_rows` here too (not
    just in fetch_export_rows) so any caller that assembles rows some other
    way (e.g. a future non-SQL source) still gets the same cap enforced.

    Never raises on empty `rows` — an empty table still produces a valid
    workbook with just the styled header row, mirroring word_export.py's
    "empty fields list does not raise" contract."""
    if len(rows) > max_rows:
        logger.info(f"build_workbook: truncating {len(rows)} rows to max_rows={max_rows}")
        rows = rows[:max_rows]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_title[:31] or "Export"  # Excel's own 31-char sheet-name limit

    header_font = Font(bold=True, color=HEADER_FONT_COLOR)
    header_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
    thin_side = Side(style=BORDER_STYLE)
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, column_name in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=str(column_name))
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = header_alignment

    for row_offset, row in enumerate(rows, start=2):
        for col_idx, column_name in enumerate(columns, start=1):
            raw_value = row[col_idx - 1] if col_idx - 1 < len(row) else None
            value = _cell_display_value(raw_value)
            cell = sheet.cell(row=row_offset, column=col_idx, value=value)
            cell.border = border
            fill = _status_fill_for(column_name, raw_value)
            if fill is not None:
                cell.fill = fill

    # Freeze the header row so it stays visible while scrolling — "A2" means
    # everything above row 2 (i.e. just row 1) is frozen.
    sheet.freeze_panes = "A2"

    # AutoFilter over the full header+data range — lets the recipient sort/
    # filter without first selecting the range themselves.
    last_column_letter = get_column_letter(max(len(columns), 1))
    last_row = max(len(rows) + 1, 1)
    sheet.auto_filter.ref = f"A1:{last_column_letter}{last_row}"

    _autosize_columns(sheet, columns, rows)

    buffer = io.BytesIO()
    workbook.save(buffer)
    data = buffer.getvalue()

    if len(data) > MAX_FILE_SIZE:
        # Logged, not raised — same "degrade, don't crash" contract as the
        # rest of this module; MAX_FILE_SIZE is a guard for CALLERS (e.g. an
        # aiohttp handler deciding whether to actually serve this) to check
        # against, not a hard stop on building it (the bytes are still valid
        # and the caller may have its own truncation/retry strategy).
        logger.warning(f"build_workbook: generated workbook is {len(data)} bytes, exceeding MAX_FILE_SIZE={MAX_FILE_SIZE}")

    logger.info(f"build_workbook: sheet_title={sheet_title!r} columns={len(columns)} rows={len(rows)} size={len(data)} bytes")
    return data
